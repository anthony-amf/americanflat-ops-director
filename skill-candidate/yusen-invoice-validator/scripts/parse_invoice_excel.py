#!/usr/bin/env python3
"""
Parse freight invoice supporting Excel files (Yusen old format and Taylored Services new format).

Extracts order numbers, quantities, and service types from the Excel files
that accompany Small Parcel/LTL invoices. Auto-detects format variants.

Usage:
    python scripts/parse_invoice_excel.py path/to/invoice.xlsx 751996              # Yusen (old)
    python scripts/parse_invoice_excel.py path/to/invoice.xlsx 752319 --warehouse FONTANA  # Taylored (new)
    python scripts/parse_invoice_excel.py path/to/invoice.xlsx INVOICE_NUMBER --output orders.json
"""

import sys
import json
import argparse
from pathlib import Path
from typing import Any

import openpyxl


def detect_format(wb) -> str:
    """Detect invoice format: 'yusen' (old) or 'taylored' (new company name)."""
    sheetnames = wb.sheetnames

    # Yusen format (old company name): has "Small Parcel" or "SML PRCL" sheet
    if any(s in sheetnames for s in ["Small Parcel", "SML PRCL", "LTL"]):
        return "yusen"

    # Taylored Services format (new company name): any sheet with a TSI PO# header
    # in A1. Sheet names vary ("Sheet1"/"Sheet2", "sp"/"ltl", ...).
    for s in sheetnames:
        first_cell = wb[s]["A1"].value
        if first_cell and "TSI" in str(first_cell).upper():
            return "taylored"

    # SHIPPED-report format: one sheet, per-carton rows under an
    # "ORDER# / CARRIER / SHIPPED / ..." header (usually row 4).
    for s in sheetnames:
        ws = wb[s]
        for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
            cells = [str(c).strip().upper() for c in row if c is not None]
            if "ORDER#" in cells and "CARRIER" in cells:
                return "shipped_report"

    return "yusen"  # Default


def parse_shipped_report_excel(file_path: str, invoice_number: str, warehouse: str = "") -> dict[str, Any]:
    """Parse the per-carton SHIPPED report (ORDER#/CARRIER/SHIPPED/HUID/UPC/UNITS/BOL/ORDER TYPE).

    One row per carton — dedupe ORDER#. Order numbers carry an AME*/AMF*/AMS*
    warehouse prefix that Stedi does NOT index; strip it (Stedi matches the bare
    ID, e.g. AMS*843F6H2W -> 843F6H2W).
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    line_items, seen = [], set()
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header_row = None
        cols = {}
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
            names = [str(c).strip().upper() if c is not None else "" for c in row]
            if "ORDER#" in names:
                header_row = i
                cols = {n: j for j, n in enumerate(names)}
                break
        if header_row is None:
            continue
        oi, ti = cols.get("ORDER#", 0), cols.get("ORDER TYPE")
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            raw = row[oi] if oi < len(row) else None
            if not raw:
                continue
            order = str(raw).strip().split("*", 1)[-1]  # strip AME*/AMF*/AMS* prefix
            if not order or order in seen:
                continue
            seen.add(order)
            otype = str(row[ti]).strip() if (ti is not None and ti < len(row) and row[ti]) else ""
            line_items.append({
                "line_item_id": len(line_items) + 1,
                "order_number": order,
                "quantity": 1,
                "service_type": "LTL" if otype.upper() == "WHOLESALE" else "Small Parcel",
                "warehouse_location": warehouse,
            })
    return {
        "invoice_number": invoice_number,
        "warehouse_location": warehouse,
        "total_line_items": len(line_items),
        "line_items": line_items,
    }


def parse_taylored_excel(file_path: str, invoice_number: str, warehouse: str = "FONTANA") -> dict[str, Any]:
    """
    Parse Taylored Services (new company name) supporting Excel file.

    Expected structure:
    - Sheet1: E-commerce/small parcel orders starting at row 14, order in column A
    - Sheet2: LTL BOLs starting at row 8, BOL in column B

    Returns:
        {
            "invoice_number": "752319",
            "warehouse": "FONTANA",
            "line_items": [
                {"order_number": "102003276483843", "service_type": "Small Parcel", "quantity": 1},
                ...
            ]
        }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    line_items = []
    seen = set()

    def add(order, stype):
        order = str(order).strip()
        if not order or order in seen:
            return
        seen.add(order)
        item = {
            "line_item_id": len(line_items) + 1,
            "order_number": order,
            "quantity": 1,
            "service_type": stype,
            "warehouse_location": warehouse,
        }
        if stype == "LTL":
            item["notes"] = "BOL number"
        line_items.append(item)

    # Header-driven: sheet names and row offsets vary across Taylored files
    # ("Sheet1"/"Sheet2" vs "sp"/"ltl"; headers at different rows). Find the
    # "Order" (parcel) or "Bol#" (LTL) header in column A of each sheet and read
    # the IDs below it, skipping blanks and continuation/total rows.
    SKIP = {"", "order", "bol#", "bol", "date", "total", "totals"}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        mode, header_row, col = None, None, 0
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            for j, c in enumerate(row[:3]):  # real headers sit in col A-C
                v = str(c).strip().lower() if c is not None else ""
                if v == "order":
                    mode, header_row, col = "Small Parcel", i, j
                    break
                if v in ("bol#", "bol"):
                    mode, header_row, col = "LTL", i, j
                    break
            if mode:
                break
        if not mode:
            continue
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            val = row[col] if row and col < len(row) else None
            if val is None or str(val).strip().lower() in SKIP:
                continue
            add(val, mode)

    if not line_items:
        # Legacy fixed-offset fallback (original Sheet1/Sheet2 files)
        if "Sheet1" in wb.sheetnames:
            for row in wb["Sheet1"].iter_rows(min_row=14, values_only=True):
                if row and row[0] and str(row[0]).strip().lower() not in SKIP:
                    add(row[0], "Small Parcel")
        if "Sheet2" in wb.sheetnames:
            for row in wb["Sheet2"].iter_rows(min_row=8, values_only=True):
                if row and len(row) > 1 and row[1]:
                    add(row[1], "LTL")

    return {
        "invoice_number": invoice_number,
        "warehouse_location": warehouse,
        "total_line_items": len(line_items),
        "line_items": line_items,
    }


def parse_yusen_excel(file_path: str, invoice_number: str, warehouse: str = "NEW JERSEY") -> dict[str, Any]:
    """
    Parse Yusen supporting Excel file.

    Expected structure:
    - "Small Parcel" sheet with order numbers in column A
    - "LTL" sheet (if present) with SSCC/BOL numbers

    Returns:
        {
            "invoice_number": "751996",
            "warehouse": "NEW JERSEY",
            "line_items": [
                {"order_number": "1Z...", "service_type": "Small Parcel", "quantity": 1},
                ...
            ]
        }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    line_items = []

    # Parse Small Parcel sheet
    if "Small Parcel" in wb.sheetnames or "SML PRCL" in wb.sheetnames:
        sheet_name = "Small Parcel" if "Small Parcel" in wb.sheetnames else "SML PRCL"
        ws = wb[sheet_name]

        # Extract order numbers from column A (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5), start=1):
            order_num = row[0].value
            quantity = row[1].value if len(row) > 1 else None
            notes = row[4].value if len(row) > 4 else None

            if order_num:  # Skip empty rows
                line_items.append({
                    "line_item_id": row_idx,
                    "order_number": str(order_num).strip(),
                    "quantity": int(quantity) if quantity else 1,
                    "service_type": "Small Parcel",
                    "warehouse_location": warehouse,
                    "notes": notes,
                })

    # Parse LTL sheet
    if "LTL" in wb.sheetnames:
        ws = wb["LTL"]

        # Extract BOL/SSCC numbers from column A
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5), start=1000):
            bol_num = row[0].value
            quantity = row[1].value if len(row) > 1 else None

            if bol_num:
                line_items.append({
                    "line_item_id": row_idx,
                    "order_number": str(bol_num).strip(),
                    "quantity": int(quantity) if quantity else 1,
                    "service_type": "LTL",
                    "warehouse_location": warehouse,
                    "notes": "BOL/SSCC number",
                })

    return {
        "invoice_number": invoice_number,
        "warehouse_location": warehouse,
        "total_line_items": len(line_items),
        "line_items": line_items,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Parse freight invoice supporting Excel files (Yusen or Taylored Services)",
        epilog="""
Examples:
  python scripts/parse_invoice_excel.py samples/751996.xlsx 751996                    # Yusen (old)
  python scripts/parse_invoice_excel.py samples/752319.xlsx 752319 --warehouse FONTANA  # Taylored (new)
  python scripts/parse_invoice_excel.py samples/invoice.xlsx INVOICE_NUMBER --output orders.json
        """
    )
    parser.add_argument("excel_file", help="Path to Excel file")
    parser.add_argument("invoice_number", help="Invoice number (e.g., 751996)")
    parser.add_argument("--warehouse", help="Warehouse location (auto-detected if not specified)")
    parser.add_argument("--output", help="Output JSON file (default: print to stdout)")

    args = parser.parse_args()

    # Verify file exists
    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        print(f"Error: File not found: {args.excel_file}", file=sys.stderr)
        sys.exit(1)

    try:
        # Detect format
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        fmt = detect_format(wb)

        # Set warehouse default based on format
        warehouse = args.warehouse or ("FONTANA" if fmt == "taylored" else "NEW JERSEY")

        # Parse with appropriate parser
        if fmt == "taylored":
            result = parse_taylored_excel(str(excel_path), args.invoice_number, warehouse)
        elif fmt == "shipped_report":
            result = parse_shipped_report_excel(str(excel_path), args.invoice_number, args.warehouse or "")
        else:
            result = parse_yusen_excel(str(excel_path), args.invoice_number, warehouse)

        # Output
        output_text = json.dumps(result, indent=2)

        if args.output:
            with open(args.output, "w") as f:
                f.write(output_text)
            print(f"✓ Parsed {result['total_line_items']} line items ({fmt}) to {args.output}")
        else:
            print(output_text)

    except Exception as e:
        print(f"Error parsing Excel: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
