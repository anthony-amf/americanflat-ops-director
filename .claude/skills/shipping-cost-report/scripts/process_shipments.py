#!/usr/bin/env python3
"""
Shipping Cost Validator for Americanflat
Matches FedEx invoices and Stamps.com print history against 3PL shipped order reports.
Outputs matched shipment details and weighted cost per unit by marketplace.
"""

import csv
import sys
import os
import re
import json
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl

# ── Marketplace normalization ──────────────────────────────────────────────

MARKETPLACE_MAP = {
    # NJ/Fontana Batch# codes
    "TARG": "Target",
    "SHOPIFY": "Shopify",
    "MICHAELS": "Michaels",
    "KOHLS": "Kohl's",
    "WAYF": "Wayfair",
    "WALC": "Walmart",
    "FAIRE": "Faire",
    "AMZC": "Amazon",
    "AMZVC": "Amazon VC",
    "S-AMZVC": "Amazon VC",
    # SC Division codes
    "TRGT": "Target",
    "SHPFY": "Shopify",
    "MCHL": "Michaels",
    "KOHL": "Kohl's",
    "FAIR": "Faire",
    "AMZCWH": "Amazon WH",
    "S-AMZCWH": "Amazon WH",
    "OTHR": "Other",
    # Additional
    "MACYS": "Macy's",
    "MACY": "Macy's",
    "SHIPSTATION": "ShipStation/Manual",
    "MANUAL": "ShipStation/Manual",
}


def normalize_marketplace(raw):
    """Normalize marketplace code to display name."""
    if not raw:
        return "Unknown"
    raw_upper = raw.strip().upper()
    raw_stripped = raw.strip()
    # Try exact match first
    if raw_stripped in MARKETPLACE_MAP:
        return MARKETPLACE_MAP[raw_stripped]
    if raw_upper in MARKETPLACE_MAP:
        return MARKETPLACE_MAP[raw_upper]
    # Try uppercase stripped
    for key, val in MARKETPLACE_MAP.items():
        if key.upper() == raw_upper:
            return val
    return raw.strip()


def clean_tracking(tracking):
    """Normalize tracking number: strip whitespace, remove Excel formatting."""
    if not tracking:
        return ""
    t = str(tracking).strip()
    # Remove Excel ="" wrapping
    t = t.replace('="', "").replace('"', "").strip()
    # Remove leading/trailing whitespace and tabs
    t = t.strip(" \t\r\n")
    return t


def parse_date(val):
    """Parse various date formats to YYYY-MM-DD string."""
    if not val:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y%m%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


# ── File parsers ───────────────────────────────────────────────────────────


def parse_fedex_invoice(filepath):
    """
    Parse FedEx invoice CSV.
    Returns dict: tracking_number -> {cost, ship_date, ref1, ref2, ref3, shipper_city}
    """
    records = {}
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            tracking = clean_tracking(
                row.get("Express or Ground Tracking ID", "")
            )
            if not tracking:
                continue
            cost = 0.0
            try:
                cost = float(row.get("Net Charge Amount", "0") or "0")
            except ValueError:
                pass
            # Weight: prefer Rated Weight (dim weight if higher), fall back to Actual
            rated_wt = 0.0
            actual_wt = 0.0
            try:
                rated_wt = float(row.get("Rated Weight Amount", "0") or "0")
            except ValueError:
                pass
            try:
                actual_wt = float(row.get("Actual Weight Amount", "0") or "0")
            except ValueError:
                pass
            weight = rated_wt if rated_wt > 0 else actual_wt

            records[tracking] = {
                "cost": cost,
                "ship_date": parse_date(row.get("Shipment Date", "")),
                "ref1": row.get("Original Customer Reference", "").strip(),
                "ref2": row.get("Original Ref#2", "").strip(),
                "ref3": row.get("Original Ref#3/PO Number", "").strip(),
                "shipper_city": row.get("Shipper City", "").strip(),
                "carrier_source": "FedEx",
                "weight": weight,
            }
    return records


def parse_stamps_history(filepath):
    """
    Parse Stamps.com print history CSV.
    Returns dict: tracking_number -> {cost, ship_date, cost_code, order_id}
    """
    records = {}
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            tracking = clean_tracking(row.get("Tracking #", ""))
            if not tracking:
                continue
            cost = 0.0
            try:
                cost = float(row.get("Amount Paid", "0") or "0")
            except ValueError:
                pass
            # Adjusted amount (refunds/adjustments)
            adjusted = 0.0
            try:
                adjusted = float(row.get("Adjusted Amount", "0") or "0")
            except ValueError:
                pass
            # Net cost = amount paid - adjusted (adjusted is the refund amount)
            net_cost = cost - adjusted if adjusted > 0 else cost

            cost_code = row.get("Cost Code", "").strip()
            order_id = row.get("Order ID", "").strip()
            printed_msg = row.get("Printed Message", "").strip()
            ref1 = row.get("Reference 1", "").strip()

            # Extract order number from multiple possible fields
            # Priority: Reference 1 > Printed Message > Cost Code
            order_ref = ""

            # Check Reference 1 (formats: "AMS*912003375746500", "AMF*...", or raw "912003352640403")
            if ref1 and ref1 != "AmerFlat123456!":
                if "*" in ref1:
                    order_ref = ref1.split("*", 1)[1].strip()
                else:
                    # Raw order number without prefix
                    order_ref = ref1

            # Check Printed Message (format: "AMF*102003366271306")
            if not order_ref and printed_msg and "*" in printed_msg:
                order_ref = printed_msg.split("*", 1)[1].strip()

            # Check Cost Code (format: "AMF*ORDERNUMBER" or "\tAMF*ORDERNUMBER")
            if not order_ref and cost_code:
                cc = cost_code.strip()
                if "*" in cc:
                    order_ref = cc.split("*", 1)[1].strip()

            # Weight
            stamps_weight = 0.0
            try:
                stamps_weight = float(row.get("Weight", "0") or "0")
            except ValueError:
                pass

            records[tracking] = {
                "cost": net_cost,
                "ship_date": parse_date(row.get("Ship Date", "")),
                "cost_code": cost_code,
                "order_id": order_id,
                "order_ref": order_ref,
                "carrier_source": "Stamps.com",
                "weight": stamps_weight,
            }
    return records


def detect_warehouse_name(filepath):
    """Detect warehouse name from filename."""
    basename = os.path.basename(filepath).lower()
    if "fontana" in basename:
        return "Fontana"
    elif "new jersey" in basename or "nj" in basename:
        return "New Jersey"
    elif "south carolina" in basename or "sc " in basename:
        return "South Carolina"
    return "Unknown"


def parse_3pl_nj_fontana(filepath, warehouse_override=None):
    """
    Parse NJ or Fontana 3PL shipped order report (CSV or XLSX).
    Returns list of dicts with standardized fields.
    """
    warehouse = warehouse_override or detect_warehouse_name(filepath)
    orders = []

    if filepath.endswith(".csv"):
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                tracking = clean_tracking(row.get("Bill of Lading", ""))
                order_num = row.get("Order", "").strip()
                po_num = row.get("PO #", "").strip()
                units = 0
                try:
                    units = int(row.get("Units", "0") or "0")
                except ValueError:
                    pass
                orders.append(
                    {
                        "order_number": order_num,
                        "po_number": po_num,
                        "tracking": tracking,
                        "ship_date": parse_date(row.get("Ship Date", "")),
                        "marketplace_raw": row.get("Batch#", "").strip(),
                        "marketplace": normalize_marketplace(
                            row.get("Batch#", "")
                        ),
                        "units": units,
                        "carrier": row.get("Carrier", "").strip(),
                        "warehouse": warehouse,
                        "consignee": row.get("Consignee", "").strip(),
                    }
                )
    else:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h).strip() if h else "" for h in row]
                continue
            if not any(row):
                continue
            vals = dict(zip(headers, row))
            tracking = clean_tracking(vals.get("Bill of Lading", ""))
            order_num = str(vals.get("Order", "")).strip()
            po_num = str(vals.get("PO #", "")).strip()
            units = 0
            try:
                units = int(vals.get("Units", 0) or 0)
            except (ValueError, TypeError):
                pass
            orders.append(
                {
                    "order_number": order_num,
                    "po_number": po_num,
                    "tracking": tracking,
                    "ship_date": parse_date(vals.get("Ship Date", "")),
                    "marketplace_raw": str(vals.get("Batch#", "")).strip()
                    if vals.get("Batch#")
                    else "",
                    "marketplace": normalize_marketplace(
                        str(vals.get("Batch#", "")).strip()
                        if vals.get("Batch#")
                        else ""
                    ),
                    "units": units,
                    "carrier": str(vals.get("Carrier", "")).strip()
                    if vals.get("Carrier")
                    else "",
                    "warehouse": warehouse,
                    "consignee": str(vals.get("Consignee", "")).strip()
                    if vals.get("Consignee")
                    else "",
                }
            )
        wb.close()
    return orders


def parse_3pl_sc(filepath):
    """
    Parse South Carolina 3PL order details (XLSX).
    This file has item-level rows; we aggregate by order+tracking to get total shipped units.
    Returns list of dicts with standardized fields.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    headers = None

    # Aggregate by (order, tracking) since SC file has item-level rows
    order_tracking_agg = {}

    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h else "" for h in row]
            continue
        if not any(row):
            continue
        vals = dict(zip(headers, row))

        tracking = clean_tracking(vals.get("Tracking Number", ""))
        order_no = str(vals.get("Order No.", "")).strip() if vals.get("Order No.") else ""
        po_no = str(vals.get("Cust PO No.", "")).strip() if vals.get("Cust PO No.") else ""

        shipped_qty = 0
        try:
            shipped_qty = int(vals.get("Shipped Quantity", 0) or 0)
        except (ValueError, TypeError):
            pass

        key = (order_no, tracking)
        if key not in order_tracking_agg:
            order_tracking_agg[key] = {
                "order_number": order_no,
                "po_number": po_no,
                "tracking": tracking,
                "ship_date": parse_date(vals.get("Actual Ship Date", "")),
                "marketplace_raw": str(vals.get("Division", "")).strip()
                if vals.get("Division")
                else "",
                "marketplace": normalize_marketplace(
                    str(vals.get("Division", "")).strip()
                    if vals.get("Division")
                    else ""
                ),
                "units": 0,
                "carrier": str(vals.get("Carrier", "")).strip()
                if vals.get("Carrier")
                else "",
                "warehouse": "South Carolina",
                "consignee": str(vals.get("Ship To Name", "")).strip()
                if vals.get("Ship To Name")
                else "",
            }
        order_tracking_agg[key]["units"] += shipped_qty

    wb.close()
    return list(order_tracking_agg.values())


# ── Matching engine ────────────────────────────────────────────────────────


def match_shipments(orders_3pl, fedex_data, stamps_data):
    """
    Match 3PL orders to FedEx/Stamps cost data.
    Returns (matched_orders, unmatched_orders, unmatched_fedex, unmatched_stamps)
    """
    matched = []
    unmatched_orders = []

    used_fedex = set()
    used_stamps = set()

    for order in orders_3pl:
        tracking = order["tracking"]
        cost_info = None
        match_method = ""

        # Primary match: tracking number
        if tracking and tracking in fedex_data:
            cost_info = fedex_data[tracking]
            used_fedex.add(tracking)
            match_method = "tracking→FedEx"
        elif tracking and tracking in stamps_data:
            cost_info = stamps_data[tracking]
            used_stamps.add(tracking)
            match_method = "tracking→Stamps"

        # Secondary match: try order/PO number against FedEx ref fields
        if not cost_info and order["po_number"]:
            po = order["po_number"].strip()
            for ft, fd in fedex_data.items():
                if ft in used_fedex:
                    continue
                if po and (fd["ref1"] == po or fd["ref2"] == po or fd["ref3"] == po):
                    cost_info = fd
                    used_fedex.add(ft)
                    match_method = "PO→FedEx ref"
                    break

        # Secondary match: try order number against Stamps order_ref
        if not cost_info and order["order_number"]:
            on = order["order_number"].strip()
            # Remove AMS* prefix if present
            on_clean = on.replace("AMS*", "")
            for st, sd in stamps_data.items():
                if st in used_stamps:
                    continue
                if sd["order_ref"] and (
                    sd["order_ref"] == on_clean or sd["order_ref"] == on
                ):
                    cost_info = sd
                    used_stamps.add(st)
                    match_method = "order→Stamps ref"
                    break

        if cost_info:
            matched.append(
                {
                    **order,
                    "shipping_cost": cost_info["cost"],
                    "carrier_source": cost_info["carrier_source"],
                    "match_method": match_method,
                    "cost_per_unit": cost_info["cost"] / max(order["units"], 1),
                    "weight": cost_info.get("weight", 0),
                }
            )
        else:
            unmatched_orders.append(order)

    # ── Reverse pass: match remaining FedEx records to 3PL orders via ref3→PO ──
    # This catches multi-package shipments where package #2+ has a different
    # tracking number but the same PO in FedEx ref fields.
    # Build a PO→order lookup from ALL 3PL orders (matched and unmatched)
    po_to_order = {}
    for order in orders_3pl:
        po = order["po_number"].strip()
        on = order["order_number"].strip()
        if po:
            po_to_order[po] = order
        if on and on != po:
            po_to_order[on] = order

    remaining_fedex = {k: v for k, v in fedex_data.items() if k not in used_fedex}
    for ft, fd in list(remaining_fedex.items()):
        ref3 = fd.get("ref3", "")
        if not ref3:
            continue
        if ref3 in po_to_order:
            base_order = po_to_order[ref3]
            matched.append(
                {
                    **base_order,
                    "shipping_cost": fd["cost"],
                    "carrier_source": "FedEx",
                    "match_method": "FedEx ref3→PO (multi-pkg)",
                    "cost_per_unit": fd["cost"] / max(base_order["units"], 1),
                    "tracking": ft,  # Use the FedEx tracking, not the 3PL BOL
                    "weight": fd.get("weight", 0),
                }
            )
            used_fedex.add(ft)

    # Same reverse pass for unmatched Stamps via order_ref
    remaining_stamps = {k: v for k, v in stamps_data.items() if k not in used_stamps}
    for st, sd in list(remaining_stamps.items()):
        ref = sd.get("order_ref", "")
        if not ref:
            continue
        # Try with and without AMS* prefix
        for lookup_key in [ref, f"AMS*{ref}"]:
            if lookup_key in po_to_order:
                base_order = po_to_order[lookup_key]
                matched.append(
                    {
                        **base_order,
                        "shipping_cost": sd["cost"],
                        "carrier_source": "Stamps.com",
                        "match_method": "Stamps ref→PO (multi-pkg)",
                        "cost_per_unit": sd["cost"] / max(base_order["units"], 1),
                        "tracking": st,
                        "weight": sd.get("weight", 0),
                    }
                )
                used_stamps.add(st)
                break

    # Identify unmatched invoice records
    unmatched_fedex = {k: v for k, v in fedex_data.items() if k not in used_fedex}
    unmatched_stamps = {k: v for k, v in stamps_data.items() if k not in used_stamps}

    return matched, unmatched_orders, unmatched_fedex, unmatched_stamps


# ── Summary calculations ──────────────────────────────────────────────────


def calculate_marketplace_summary(matched_orders):
    """
    Calculate weighted cost per unit by marketplace.
    Returns list of dicts sorted by marketplace.
    """
    mp_data = defaultdict(lambda: {"total_cost": 0.0, "total_units": 0, "order_count": 0})

    for order in matched_orders:
        mp = order["marketplace"]
        mp_data[mp]["total_cost"] += order["shipping_cost"]
        mp_data[mp]["total_units"] += order["units"]
        mp_data[mp]["order_count"] += 1

    summary = []
    for mp, data in sorted(mp_data.items()):
        weighted_cpu = data["total_cost"] / max(data["total_units"], 1)
        summary.append(
            {
                "marketplace": mp,
                "total_orders": data["order_count"],
                "total_units": data["total_units"],
                "total_shipping_cost": round(data["total_cost"], 2),
                "weighted_cost_per_unit": round(weighted_cpu, 2),
                "avg_cost_per_order": round(
                    data["total_cost"] / max(data["order_count"], 1), 2
                ),
            }
        )
    return summary


def calculate_warehouse_summary(matched_orders):
    """Breakdown by warehouse."""
    wh_data = defaultdict(lambda: {"total_cost": 0.0, "total_units": 0, "order_count": 0})
    for order in matched_orders:
        wh = order["warehouse"]
        wh_data[wh]["total_cost"] += order["shipping_cost"]
        wh_data[wh]["total_units"] += order["units"]
        wh_data[wh]["order_count"] += 1

    summary = []
    for wh, data in sorted(wh_data.items()):
        summary.append(
            {
                "warehouse": wh,
                "total_orders": data["order_count"],
                "total_units": data["total_units"],
                "total_shipping_cost": round(data["total_cost"], 2),
                "weighted_cost_per_unit": round(
                    data["total_cost"] / max(data["total_units"], 1), 2
                ),
            }
        )
    return summary


# ── Excel output ───────────────────────────────────────────────────────────


def write_excel_report(
    matched,
    unmatched_orders,
    marketplace_summary,
    warehouse_summary,
    unmatched_fedex,
    unmatched_stamps,
    output_path,
):
    """Write comprehensive Excel report with multiple sheets."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Marketplace Summary ──
    ws1 = wb.active
    ws1.title = "Marketplace Summary"
    headers1 = [
        "Marketplace",
        "Total Orders",
        "Total Units",
        "Total Shipping Cost",
        "Weighted Cost/Unit",
        "Avg Cost/Order",
    ]
    ws1.append(headers1)
    for row in marketplace_summary:
        ws1.append(
            [
                row["marketplace"],
                row["total_orders"],
                row["total_units"],
                row["total_shipping_cost"],
                row["weighted_cost_per_unit"],
                row["avg_cost_per_order"],
            ]
        )
    # Totals row
    total_orders = sum(r["total_orders"] for r in marketplace_summary)
    total_units = sum(r["total_units"] for r in marketplace_summary)
    total_cost = sum(r["total_shipping_cost"] for r in marketplace_summary)
    ws1.append(
        [
            "TOTAL",
            total_orders,
            total_units,
            round(total_cost, 2),
            round(total_cost / max(total_units, 1), 2),
            round(total_cost / max(total_orders, 1), 2),
        ]
    )

    # Format currency columns
    for row in ws1.iter_rows(min_row=2, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = '$#,##0.00'

    # Bold headers and totals
    from openpyxl.styles import Font, PatternFill

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    for cell in ws1[1]:
        cell.font = header_font_white
        cell.fill = header_fill
    for cell in ws1[ws1.max_row]:
        cell.font = header_font

    # Auto-width
    for col in ws1.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 2
        ws1.column_dimensions[col[0].column_letter].width = min(max_len, 30)

    # ── Sheet 2: Warehouse Summary ──
    ws2 = wb.create_sheet("Warehouse Summary")
    headers2 = [
        "Warehouse",
        "Total Orders",
        "Total Units",
        "Total Shipping Cost",
        "Weighted Cost/Unit",
    ]
    ws2.append(headers2)
    for row in warehouse_summary:
        ws2.append(
            [
                row["warehouse"],
                row["total_orders"],
                row["total_units"],
                row["total_shipping_cost"],
                row["weighted_cost_per_unit"],
            ]
        )
    for cell in ws2[1]:
        cell.font = header_font_white
        cell.fill = header_fill
    for row in ws2.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = '$#,##0.00'
    for col in ws2.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 2
        ws2.column_dimensions[col[0].column_letter].width = min(max_len, 30)

    # ── Sheet 3: Matched Orders Detail ──
    ws3 = wb.create_sheet("Matched Orders")
    headers3 = [
        "Ship Date",
        "Order Number",
        "PO #",
        "Marketplace",
        "Units",
        "Shipping Cost",
        "Cost/Unit",
        "Carrier Source",
        "Match Method",
        "Warehouse",
        "Tracking",
        "Consignee",
    ]
    ws3.append(headers3)
    for order in sorted(matched, key=lambda x: (x["marketplace"], x.get("ship_date", ""))):
        ws3.append(
            [
                order["ship_date"],
                order["order_number"],
                order["po_number"],
                order["marketplace"],
                order["units"],
                order["shipping_cost"],
                round(order["cost_per_unit"], 2),
                order["carrier_source"],
                order["match_method"],
                order["warehouse"],
                order["tracking"],
                order["consignee"],
            ]
        )
    for cell in ws3[1]:
        cell.font = header_font_white
        cell.fill = header_fill
    for row in ws3.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = '$#,##0.00'
    for col in ws3.columns:
        max_len = max(min(len(str(c.value or "")) for c in col if c.value) + 5, 12)
        ws3.column_dimensions[col[0].column_letter].width = min(max_len, 30)

    # ── Sheet 4: Unmatched 3PL Orders ──
    ws4 = wb.create_sheet("Unmatched 3PL Orders")
    headers4 = [
        "Order Number",
        "PO #",
        "Marketplace",
        "Units",
        "Ship Date",
        "Carrier",
        "Tracking",
        "Warehouse",
        "Consignee",
    ]
    ws4.append(headers4)
    for order in sorted(unmatched_orders, key=lambda x: x.get("marketplace", "")):
        ws4.append(
            [
                order["order_number"],
                order["po_number"],
                order["marketplace"],
                order["units"],
                order["ship_date"],
                order["carrier"],
                order["tracking"],
                order["warehouse"],
                order["consignee"],
            ]
        )
    for cell in ws4[1]:
        cell.font = header_font_white
        cell.fill = header_fill
    for col in ws4.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 2
        ws4.column_dimensions[col[0].column_letter].width = min(max_len, 30)

    # ── Sheet 5: Unmatched Invoice Records ──
    ws5 = wb.create_sheet("Unmatched Invoices")
    headers5 = ["Source", "Tracking", "Cost", "Ship Date", "Reference"]
    ws5.append(headers5)
    for tracking, data in sorted(unmatched_fedex.items()):
        ws5.append(
            [
                "FedEx",
                tracking,
                data["cost"],
                data["ship_date"],
                f"Ref1: {data['ref1']} | Ref3: {data['ref3']}",
            ]
        )
    for tracking, data in sorted(unmatched_stamps.items()):
        ws5.append(
            [
                "Stamps.com",
                tracking,
                data["cost"],
                data["ship_date"],
                f"Order ref: {data.get('order_ref', '')}",
            ]
        )
    for cell in ws5[1]:
        cell.font = header_font_white
        cell.fill = header_fill
    for row in ws5.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '$#,##0.00'
    for col in ws5.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 2
        ws5.column_dimensions[col[0].column_letter].width = min(max_len, 40)

    wb.save(output_path)
    return output_path


# ── Google Sheets JSON output ──────────────────────────────────────────────


def write_gsheets_json(marketplace_summary, date_range, output_path):
    """
    Write marketplace summary as JSON for appending to Google Sheets.
    Includes date range metadata for the running record.
    """
    output = {
        "date_range": date_range,
        "generated_at": datetime.now().isoformat(),
        "marketplace_data": marketplace_summary,
    }
    with open(output_path, "w") as f:
        json.dump(output, f, indent=2)
    return output_path


# ── Main entry point ───────────────────────────────────────────────────────


def run(
    fedex_files=None,
    stamps_files=None,
    nj_fontana_files=None,
    sc_files=None,
    output_dir=".",
):
    """
    Main processing function.
    All parameters are lists of file paths.
    Returns dict with summary stats and output file paths.
    """
    fedex_files = fedex_files or []
    stamps_files = stamps_files or []
    nj_fontana_files = nj_fontana_files or []
    sc_files = sc_files or []

    # Parse all invoice/cost sources
    print("Parsing FedEx invoices...")
    fedex_data = {}
    for f in fedex_files:
        data = parse_fedex_invoice(f)
        fedex_data.update(data)
        print(f"  {os.path.basename(f)}: {len(data)} records")

    print("Parsing Stamps.com print history...")
    stamps_data = {}
    for f in stamps_files:
        data = parse_stamps_history(f)
        stamps_data.update(data)
        print(f"  {os.path.basename(f)}: {len(data)} records")

    # Parse all 3PL reports
    print("Parsing 3PL shipped order reports...")
    all_orders = []
    for f in nj_fontana_files:
        orders = parse_3pl_nj_fontana(f)
        all_orders.extend(orders)
        print(f"  {os.path.basename(f)}: {len(orders)} orders")

    for f in sc_files:
        orders = parse_3pl_sc(f)
        all_orders.extend(orders)
        print(f"  {os.path.basename(f)}: {len(orders)} orders")

    print(f"\nTotal 3PL orders: {len(all_orders)}")
    print(f"Total FedEx records: {len(fedex_data)}")
    print(f"Total Stamps records: {len(stamps_data)}")

    # Match
    print("\nMatching shipments...")
    matched, unmatched_orders, unmatched_fedex, unmatched_stamps = match_shipments(
        all_orders, fedex_data, stamps_data
    )

    print(f"  Matched: {len(matched)}")
    print(f"  Unmatched 3PL orders: {len(unmatched_orders)}")
    print(f"  Unmatched FedEx records: {len(unmatched_fedex)}")
    print(f"  Unmatched Stamps records: {len(unmatched_stamps)}")

    # Match method breakdown
    method_counts = defaultdict(int)
    for m in matched:
        method_counts[m["match_method"]] += 1
    print("\n  Match methods:")
    for method, count in sorted(method_counts.items()):
        print(f"    {method}: {count}")

    # Calculate summaries
    marketplace_summary = calculate_marketplace_summary(matched)
    warehouse_summary = calculate_warehouse_summary(matched)

    # Determine date range from matched orders
    dates = [m["ship_date"] for m in matched if m["ship_date"]]
    date_range = ""
    if dates:
        date_range = f"{min(dates)} to {max(dates)}"

    # Write outputs
    print("\nWriting Excel report...")
    excel_path = os.path.join(output_dir, "shipping_cost_report.xlsx")
    write_excel_report(
        matched,
        unmatched_orders,
        marketplace_summary,
        warehouse_summary,
        unmatched_fedex,
        unmatched_stamps,
        excel_path,
    )
    print(f"  Saved: {excel_path}")

    # Write Google Sheets JSON
    json_path = os.path.join(output_dir, "marketplace_summary.json")
    write_gsheets_json(marketplace_summary, date_range, json_path)
    print(f"  Saved: {json_path}")

    # Print summary table
    print("\n" + "=" * 80)
    print("MARKETPLACE WEIGHTED COST PER UNIT")
    print("=" * 80)
    print(
        f"{'Marketplace':<20} {'Orders':>8} {'Units':>8} {'Total Cost':>12} {'Cost/Unit':>10}"
    )
    print("-" * 60)
    for row in marketplace_summary:
        print(
            f"{row['marketplace']:<20} {row['total_orders']:>8} {row['total_units']:>8} "
            f"${row['total_shipping_cost']:>10,.2f} ${row['weighted_cost_per_unit']:>8,.2f}"
        )
    total_orders = sum(r["total_orders"] for r in marketplace_summary)
    total_units = sum(r["total_units"] for r in marketplace_summary)
    total_cost = sum(r["total_shipping_cost"] for r in marketplace_summary)
    print("-" * 60)
    print(
        f"{'TOTAL':<20} {total_orders:>8} {total_units:>8} "
        f"${total_cost:>10,.2f} ${total_cost / max(total_units, 1):>8,.2f}"
    )

    return {
        "matched_count": len(matched),
        "unmatched_3pl_count": len(unmatched_orders),
        "unmatched_fedex_count": len(unmatched_fedex),
        "unmatched_stamps_count": len(unmatched_stamps),
        "marketplace_summary": marketplace_summary,
        "warehouse_summary": warehouse_summary,
        "date_range": date_range,
        "excel_path": excel_path,
        "json_path": json_path,
    }


if __name__ == "__main__":
    # CLI usage for testing
    import argparse

    parser = argparse.ArgumentParser(description="Shipping Cost Validator")
    parser.add_argument("--fedex", nargs="+", help="FedEx invoice CSV files")
    parser.add_argument("--stamps", nargs="+", help="Stamps.com print history CSV files")
    parser.add_argument(
        "--nj-fontana", nargs="+", help="NJ/Fontana 3PL shipped order reports"
    )
    parser.add_argument("--sc", nargs="+", help="South Carolina 3PL order details")
    parser.add_argument(
        "--output-dir", default=".", help="Output directory for reports"
    )

    args = parser.parse_args()

    result = run(
        fedex_files=args.fedex or [],
        stamps_files=args.stamps or [],
        nj_fontana_files=args.nj_fontana or [],
        sc_files=args.sc or [],
        output_dir=args.output_dir,
    )
