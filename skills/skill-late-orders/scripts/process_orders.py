#!/usr/bin/env python3
"""
AMF Late-Order processor — reusable engine for the amf-late-orders skill.

Implements the full SLA ruleset described in SKILL.md:
  * Mon–Fri processing at all three warehouses (Sat + Sun non-processing).
  * Two-step weekend-receipt deadline:
        effective receipt = driving date rolled forward to a processing day,
        deadline          = next processing day AFTER the effective receipt.
  * AMZC (Amazon DF) orders are exempt from the weekend logic and run purely
    on CANCELDATE.
  * Standard format (NJ / Fontana): drives off RF.DATE, drops AMZVC / S-AMZVC,
    has a PENDING ACKNOWLEDGEMENT bucket (blank RF.DATE on a non-AMZC order).
  * SC variant: drives off Order Date, keeps Order Type == 'ECOM Order',
    de-dups on Order No., drops amazon.com wholesale (AMZCWH / S-AMZCWH),
    no pending-ack bucket.

This script is a STARTING POINT. Standing exclusions, one-time exclusions,
holidays, and cancel lists change run to run — read SKILL.md and adjust the
constants / call arguments below for the day you are processing. Always
confirm the run timestamp (used as "now") from the uploaded filename.

Usage (from the skill, inside the code tool):
    from process_orders import process_standard, process_sc
    res = process_standard("/path/Order_Report_YYYYMMDD_HHMMSS.xlsx",
                            short="NJ", full="New Jersey",
                            ts="YYYYMMDDHHMMSS",
                            standing_exclude={"AME*6-05-2026", ...},
                            one_time_exclude=set(),
                            holidays=set())
"""

from __future__ import annotations
import os
from datetime import datetime, timedelta, date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "/mnt/user-data/outputs"

STATUS_STYLES = {
    "PAST DUE":                ("FFC7CE", "9C0006"),
    "SHIP BY EOD":             ("FFEB9C", "9C5700"),
    "PENDING ACKNOWLEDGEMENT": ("D9D9D9", "595959"),
}

# Batch codes that are known / in-scope (anything else gets flagged for review).
KNOWN_BATCHES = {
    "AMZC", "TARG", "WALC", "WAYF", "SHOPIFY", "MACY",
    "MICHAELS", "KOHLS", "FAIRE", "OTHR", "N",
}

# Amazon Vendor Central — always dropped from the standard format.
VENDOR_CENTRAL = {"AMZVC", "S-AMZVC"}

# Amazon.com wholesale — always dropped from the SC format.
SC_AMAZON_WHOLESALE = {"AMZCWH", "S-AMZCWH"}

# Standing order-ID exclusions, applied automatically every run until cleared.
# These are specific orders handled outside the normal flow. Remove an entry
# only on the operator's explicit instruction. Keyed by warehouse short code.
STANDING_EXCLUSIONS = {
    "NJ": {
        "AME*6-05-2026",       # manual order, handled separately (added 06/16/26)
        "AME*SSXDAUTEWF6H2",   # manual order, handled separately (added 06/16/26)
        "AME*MUSA00210333",    # bulk buy, no ship-window date (added 06/25/26)
        "AME*VUXFU52ETJXFY",   # wholesale, no definitive ship date (added 06/30/26)
        "AME*CG062526-NJ",     # recurring stuck order (added 07/17/26)
    },
    "Fontana": {
        "AMF*MUSA00210332",    # bulk buy, no ship-window date (added 06/25/26)
        "AMF*PxT78SYST",       # stuck AMZC order, force-cancel pending (added 07/14/26)
        "AMF*CG062526-FON",    # recurring stuck order (added 07/14/26)
    },
    "SC": {
        "AMS*CG062526-SC",     # recurring stuck order (added 07/17/26)
    },
}


# ----------------------------------------------------------------------------
# Date helpers
# ----------------------------------------------------------------------------
def parse_date(s):
    if pd.isna(s) or s == "":
        return None
    try:
        return datetime.strptime(str(s), "%m/%d/%Y").date()
    except ValueError:
        return None


def fmt_date_cell(d):
    if pd.isna(d):
        return ""
    if hasattr(d, "date"):
        d = d.date()
    return d.strftime("%m/%d/%Y") if hasattr(d, "strftime") else str(d)


def make_calendar(holidays: set[date] | None):
    """Return (is_nonproc, roll_to_processing, npd, deadline_for) closures.

    All warehouses are Mon–Fri. Saturday (weekday 5) and Sunday (weekday 6)
    are non-processing, plus any dates in `holidays`.
    """
    holidays = holidays or set()

    def is_nonproc(d: date) -> bool:
        return d in holidays or d.weekday() in (5, 6)

    def roll_to_processing(d: date) -> date:
        while is_nonproc(d):
            d += timedelta(days=1)
        return d

    def npd(d: date) -> date:
        nd = d + timedelta(days=1)
        while is_nonproc(nd):
            nd += timedelta(days=1)
        return nd

    def deadline_for(driving: date) -> date:
        # Two-step: roll weekend/holiday receipt to a working day, then +1 processing day.
        return npd(roll_to_processing(driving))

    return is_nonproc, roll_to_processing, npd, deadline_for


# ----------------------------------------------------------------------------
# Email builder
# ----------------------------------------------------------------------------
def build_email(full_name, past_ids, ship_count, pending_count, today, yesterday,
                is_sc=False, se_ids=None, cancel_ids=None, df_ids=None):
    """Return (subject, body) following the SKILL.md email rules.

    Past-due-only list by default. SC falls back to listing ship-by-EOD POs
    when there are zero past due (pass se_ids for that). cancel_ids adds a
    one-off "please CANCEL" block when supplied. df_ids are the AMZC orders
    whose CANCELDATE is today — they get their own itemised block, and on
    Fontana only they also add a count segment to the subject.
    """
    M, N = len(past_ids), len(past_ids) + ship_count
    df_ids = list(df_ids or [])

    if M > 0 and ship_count > 0:
        subj = f"AMF x TS {full_name} {N} POs to Ship by EOD {today:%m/%d/%y} --- {M} POs Past Due"
    elif M == 0 and ship_count > 0:
        subj = f"AMF x TS {full_name} {N} POs to Ship by EOD {today:%m/%d/%y}"
    elif M > 0:
        subj = f"AMF x TS {full_name} {M} POs Past Due {today:%m/%d/%y}"
    else:
        subj = f"AMF x TS {full_name} all clear {today:%m/%d/%y}"

    # Amazon DF ship-today count — Fontana subject only.
    if full_name == "Fontana" and df_ids:
        subj += f" --- {len(df_ids)} DF to Ship Today"

    if M == 0 and ship_count == 0 and pending_count == 0 and not cancel_ids:
        body = (f"Good morning team — open order review for {today:%m/%d/%y} "
                f"is all clear. No late or at-risk orders today.")
        return subj, body

    lines = ["Good morning team", ""]
    if M > 0 and ship_count > 0:
        lines.append(f"I am seeing {N} x POs that must ship by EOD to avoid a late flag. "
                     f"Note that {M} of these are past due and MUST ship today!")
    elif M == 0 and ship_count > 0:
        lines.append(f"I am seeing {N} x POs that must ship by EOD to avoid a late flag. "
                     f"None are past due yet — let's keep it that way.")
    elif M > 0:
        lines.append(f"I am seeing {M} x POs that are past due and MUST ship today!")
    else:
        lines.append("No late or at-risk orders today.")
    lines.append("")

    if M > 0:
        lines.append("Past Due below, full list in report.")
        for oid in past_ids[:25]:
            lines.append(oid)
        if len(past_ids) > 25:
            lines.append(f"... and {len(past_ids) - 25} more — see attached spreadsheet.")
        lines.append("")
    elif is_sc and ship_count > 0 and se_ids:
        # SC-only fallback so the email isn't empty on a light, no-past-due day.
        lines.append("POs to ship by EOD below, full list in report.")
        for oid in se_ids[:25]:
            lines.append(oid)
        if len(se_ids) > 25:
            lines.append(f"... and {len(se_ids) - 25} more — see attached spreadsheet.")
        lines.append("")

    if df_ids:
        lines.append("Amazon DF orders with a cancel date of TODAY — these MUST ship "
                     "today or Amazon cancels them:")
        for oid in df_ids[:25]:
            lines.append(oid)
        if len(df_ids) > 25:
            lines.append(f"... and {len(df_ids) - 25} more — see attached spreadsheet.")
        lines.append("")

    if cancel_ids:
        lines.append("Also, please CANCEL the following orders — do NOT ship these, "
                     "they have been cancelled:")
        for oid in cancel_ids:
            lines.append(oid)
        lines.append("")

    if not is_sc and pending_count > 0:
        unit = "order" if pending_count == 1 else "orders"
        verb = "has" if pending_count == 1 else "have"
        lines.append(f"We also have {pending_count} {unit} that {verb} yet to be "
                     f"acknowledged, please review and process once time permits.")
        lines.append("")

    lines.append(f"Please see full report for full list of POs. All highlighted POs either "
                 f"have ship dates for today or are open orders from {yesterday:%m/%d/%y} or "
                 f"earlier that are late or will be considered late if not shipped by EOD.")
    return subj, "\n".join(lines)


# ----------------------------------------------------------------------------
# XLSX writer
# ----------------------------------------------------------------------------
def _write_xlsx(rows, columns, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Flagged Orders"
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", start_color="E7E6E6")
        cell.font = Font(bold=True, name="Calibri")
        cell.alignment = Alignment(horizontal="left")
    for status, values in rows:
        ws.append(values)
        fill_hex, font_hex = STATUS_STYLES[status]
        sc = ws.cell(row=ws.max_row, column=1)
        sc.fill = PatternFill("solid", start_color=fill_hex)
        sc.font = Font(bold=True, color=font_hex, name="Calibri")
    ws.freeze_panes = "A2"
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)
    for r in range(2, ws.max_row + 1):
        for c in range(2, len(columns) + 1):
            ws.cell(row=r, column=c).font = Font(name="Calibri")
    if os.path.exists(out_path):
        os.remove(out_path)
    wb.save(out_path)


# ----------------------------------------------------------------------------
# Standard format (New Jersey, Fontana)
# ----------------------------------------------------------------------------
def process_standard(path, short, full, ts,
                     standing_exclude=None, one_time_exclude=None,
                     holidays=None, cancel_ids=None):
    # Baked-in standing exclusions for this warehouse are always applied.
    # Anything passed via standing_exclude is merged on top (additive).
    standing_exclude = set(standing_exclude or set()) | STANDING_EXCLUSIONS.get(short, set())
    one_time_exclude = one_time_exclude or set()
    cancel_ids = list(cancel_ids or [])
    _, _, _, deadline_for = make_calendar(holidays)

    now = datetime.strptime(ts, "%Y%m%d%H%M%S")
    today = now.date()
    yesterday = today - timedelta(days=1)

    df = pd.read_excel(path)
    total = len(df)
    df = df[~df["BATCHNO"].isin(VENDOR_CENTRAL)].copy()
    excl = standing_exclude | one_time_exclude | {c.lower() for c in cancel_ids}
    df = df[~df["ORDERID"].astype(str).str.lower().isin({e.lower() for e in excl})].copy()
    in_scope = len(df)

    df["_cd"] = df["CANCELDATE"].apply(parse_date)
    df["_rf"] = df["RF.DATE"].apply(parse_date)
    df["_ed"] = df["ENTRYDATE"].apply(parse_date)
    df["_sd"] = df["STARTDATE"].apply(parse_date) if "STARTDATE" in df.columns else None

    def assign(r):
        # Target runs on STARTDATE (its Requested Ship Date), not RF.DATE + 1.
        # No weekend/holiday roll-forward: Target's ~10:00 placement cutoff is
        # already baked into the RSD, so rolling would under-report lateness.
        if r["BATCHNO"] == "TARG" and r.get("_sd") is not None:
            sd = r["_sd"]
            if sd < today:
                return sd, "PAST DUE"
            if sd == today:
                return sd, "SHIP BY EOD"
            return sd, "EXCLUDED"
        if r["BATCHNO"] == "AMZC":
            cd = r["_cd"]
            if cd is None:
                return None, "DATA ISSUE"
            if cd < today:
                return cd, "PAST DUE"
            if cd == today:
                return cd, "SHIP BY EOD"
            return cd, "EXCLUDED"
        rd = r["_rf"]
        if rd is None:
            return None, "PENDING ACKNOWLEDGEMENT"
        dl = deadline_for(rd)
        if dl < today:
            return dl, "PAST DUE"
        if dl == today:
            return dl, "SHIP BY EOD"
        return dl, "EXCLUDED"

    res = df.apply(assign, axis=1)
    df["DEADLINE"] = [r[0] for r in res]
    df["STATUS"] = [r[1] for r in res]
    df["DAYS_LATE"] = df.apply(
        lambda r: (today - r["DEADLINE"]).days
        if r["DEADLINE"] and r["STATUS"] == "PAST DUE" else None, axis=1)

    flagged = df[df["STATUS"].isin(["PAST DUE", "SHIP BY EOD", "PENDING ACKNOWLEDGEMENT"])].copy()
    past = flagged[flagged["STATUS"] == "PAST DUE"].sort_values(
        ["DAYS_LATE", "ORDERID"], ascending=[False, True])
    se = flagged[flagged["STATUS"] == "SHIP BY EOD"].copy()
    se["_a"] = (se["BATCHNO"] != "AMZC").astype(int)
    se = se.sort_values(["_a", "_rf", "ORDERID"], na_position="last")
    pend = flagged[flagged["STATUS"] == "PENDING ACKNOWLEDGEMENT"].sort_values(
        "_ed", na_position="last")
    ordered = pd.concat([past, se, pend], ignore_index=True)

    columns = ["STATUS", "ORDERID", "BATCHNO", "ORDERSTATUS", "CONSNAME", "PONUM",
               "ENTRYDATE", "RF.DATE", "ENTRY-RF LAG", "RF.TIME", "CANCELDATE",
               "SHIPVIA", "UNITS", "NUM.OF.LINES", "DEADLINE", "DAYS LATE"]
    rows = []
    for _, row in ordered.iterrows():
        deadline_str = row["DEADLINE"].strftime("%m/%d/%Y") if row["DEADLINE"] else ""
        if row["STATUS"] == "PAST DUE":
            dl = int(row["DAYS_LATE"]) if row["DAYS_LATE"] is not None else ""
        elif row["STATUS"] == "SHIP BY EOD":
            dl = 0
        else:
            dl = ""
        # Acknowledgement latency: how long the order sat before it was RF'd.
        # Blank when either date is missing (always blank for PENDING ACK).
        lag = (row["_rf"] - row["_ed"]).days if (row["_rf"] and row["_ed"]) else ""
        rows.append((row["STATUS"], [
            row["STATUS"], row["ORDERID"],
            row["BATCHNO"] if pd.notna(row["BATCHNO"]) else "",
            row["ORDERSTATUS"],
            row["CONSNAME"] if pd.notna(row["CONSNAME"]) else "",
            row["PONUM"] if pd.notna(row["PONUM"]) else "",
            row["ENTRYDATE"] if pd.notna(row["ENTRYDATE"]) else "",
            row["RF.DATE"] if pd.notna(row["RF.DATE"]) else "",
            lag,
            row["RF.TIME"] if pd.notna(row["RF.TIME"]) else "",
            row["CANCELDATE"] if pd.notna(row["CANCELDATE"]) else "",
            row["SHIPVIA"] if pd.notna(row["SHIPVIA"]) else "",
            int(row["UNITS"]) if pd.notna(row["UNITS"]) else "",
            int(row["NUM.OF.LINES"]) if pd.notna(row["NUM.OF.LINES"]) else "",
            deadline_str, dl,
        ]))

    out_path = os.path.join(OUTPUT_DIR, f"{short}_Flagged_Orders_{now:%Y%m%d_%H%M%S}.xlsx")
    _write_xlsx(rows, columns, out_path)

    unknown = set(df["BATCHNO"].dropna().unique()) - KNOWN_BATCHES
    # Amazon DF orders whose cancel date is today — itemised in the email so the
    # warehouse can see exactly which orders auto-cancel if they don't go out.
    df_today = sorted(se[(se["BATCHNO"] == "AMZC") & (se["_cd"] == today)]["ORDERID"].tolist())
    subject, body = build_email(full, past["ORDERID"].tolist(), len(se), len(pend),
                                today, yesterday, is_sc=False, cancel_ids=cancel_ids,
                                df_ids=df_today)
    return {
        "warehouse": full, "today": today, "total": total, "in_scope": in_scope,
        "past_due": len(past), "ship_by_eod": len(se), "pending": len(pend),
        "unknown_batches": unknown, "path": out_path,
        "subject": subject, "body": body, "frame": df,
    }


# ----------------------------------------------------------------------------
# SC variant (South Carolina)
# ----------------------------------------------------------------------------
def process_sc(path, ts, one_time_exclude=None, holidays=None,
               drop_amazon_wholesale=True, cancel_ids=None):
    one_time_exclude = one_time_exclude or set()
    cancel_ids = list(cancel_ids or [])
    _, _, _, deadline_for = make_calendar(holidays)

    now = datetime.strptime(ts, "%Y%m%d%H%M%S")
    today = now.date()
    yesterday = today - timedelta(days=1)

    df = pd.read_excel(path)
    total = len(df)
    if drop_amazon_wholesale:
        df = df[~df["Division"].astype(str).str.strip().isin(SC_AMAZON_WHOLESALE)].copy()
    df = df[df["Order Type"] == "ECOM Order"].copy()
    df = df.drop_duplicates(subset=["Order No."], keep="first").copy()
    excl = {e.lower() for e in (one_time_exclude | {c.lower() for c in cancel_ids})}
    df = df[~df["Order No."].astype(str).str.lower().isin(excl)].copy()
    in_scope = len(df)

    def cat(od):
        if pd.isna(od):
            return None, "NO ORDER DATE"
        d = od.date()
        dl = deadline_for(d)
        if dl < today:
            return d, "PAST DUE"
        if dl == today:
            return d, "SHIP BY EOD"
        return d, "EXCLUDED"

    res = df["Order Date"].apply(cat)
    df["_od"] = [r[0] for r in res]
    df["STATUS"] = [r[1] for r in res]
    df["DAYS_LATE"] = df.apply(
        lambda r: (today - deadline_for(r["_od"])).days
        if r["STATUS"] == "PAST DUE" else None, axis=1)

    flagged = df[df["STATUS"].isin(["PAST DUE", "SHIP BY EOD"])].copy()
    past = flagged[flagged["STATUS"] == "PAST DUE"].sort_values(
        ["DAYS_LATE", "Order No."], ascending=[False, True])
    se = flagged[flagged["STATUS"] == "SHIP BY EOD"].sort_values("Order No.")
    ordered = pd.concat([past, se], ignore_index=True)

    columns = ["STATUS", "Order No.", "Division", "Ship To", "Cust PO No.",
               "Order Date", "Earliest Ship Date", "Latest Ship Date", "Carrier",
               "Total Lines", "Total Units", "DAYS LATE"]
    rows = []
    for _, row in ordered.iterrows():
        dl = int(row["DAYS_LATE"]) if row["STATUS"] == "PAST DUE" else 0
        rows.append((row["STATUS"], [
            row["STATUS"], row["Order No."],
            row["Division"].strip() if pd.notna(row["Division"]) else "",
            row["Ship To"] if pd.notna(row["Ship To"]) else "",
            row["Cust PO No."] if pd.notna(row["Cust PO No."]) else "",
            fmt_date_cell(row["Order Date"]),
            fmt_date_cell(row["Earliest Ship Date"]),
            fmt_date_cell(row["Latest Ship Date"]),
            row["Carrier"] if pd.notna(row["Carrier"]) else "",
            int(row["Total Lines"]) if pd.notna(row["Total Lines"]) else "",
            int(row["Total Units"]) if pd.notna(row["Total Units"]) else "",
            dl,
        ]))

    out_path = os.path.join(OUTPUT_DIR, f"SC_Flagged_Orders_{now:%Y%m%d_%H%M%S}.xlsx")
    _write_xlsx(rows, columns, out_path)

    subject, body = build_email("South Carolina", past["Order No."].tolist(), len(se),
                                0, today, yesterday, is_sc=True,
                                se_ids=se["Order No."].tolist(), cancel_ids=cancel_ids)
    return {
        "warehouse": "South Carolina", "today": today, "total": total, "in_scope": in_scope,
        "past_due": len(past), "ship_by_eod": len(se), "pending": 0,
        "path": out_path, "subject": subject, "body": body, "frame": df,
    }
