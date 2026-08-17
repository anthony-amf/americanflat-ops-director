#!/usr/bin/env python3
"""Match Amazon Vendor Central open payables against the remittance detail.

Two exports from Vendor Central go in:

  1. "Your payables - ready for deduction"  (CSV)
     Every open chargeback / product return / contra-COGS item Amazon still
     intends to net out of a future payment.

  2. "Payments"  (XLSX, sheet "Remittance payments")
     Two stacked sections: a payment header list, then an "Invoices" section
     with the line-by-line detail of what each payment actually settled.

The join key is the payables "Transaction number" == the remittance
"Invoice Number". Both use the same identifiers (1520811216VCBSINV,
6268-1490026100, 365942281678551, ...).

Output is an Excel workbook with one tab per bucket so AP can work the list:

  Summary                  headline numbers and the tie-out
  1 Already deducted       open payables that a remittance already settled
                           -- these overstate the open balance
  2 Still open             open payables with no remittance hit in the window
  3 Deducted not on list   remittance deductions absent from the payables file
  4 Payment tie-out        per payment: header amount vs sum of its lines
  5 Multi-payment claims   claims Amazon applied in slices across payments

Usage:
    python3 amazon-payables-remittance-match.py PAYABLES.csv PAYMENTS.xlsx \
        [--out amazon-payables-remittance-match.xlsx]

Important scope note: the remittance export only covers the payments it
contains. A payable that shows up in "2 Still open" is only open *as far as
this export can see* -- if it was deducted before the earliest payment in the
file, there is no line here to match it to. The Summary tab states the window.
"""

import argparse
import csv
import datetime as dt
from collections import defaultdict
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MONEY = '#,##0.00;[Red]-#,##0.00'

# Header rows in the payments workbook are found by these labels, so the
# script keeps working if Amazon shifts the sections up or down a row.
PAYMENT_HDR = 'Payment Number'
INVOICE_HDR_MARKER = 'Invoice Number'


def money(value):
    """Parse '$1,234.56', '(12.30)', '-12.30', 12.3 -> Decimal. None if blank."""
    if value is None or value == '':
        return None
    text = str(value).strip().replace('$', '').replace(',', '')
    negative = text.startswith('(')
    text = text.strip('()')
    if not text:
        return None
    try:
        amount = Decimal(text)
    except Exception:
        return None
    return -amount if negative else amount


def parse_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    for fmt in ('%m/%d/%Y', '%Y-%m-%d'):
        try:
            return dt.datetime.strptime(str(value).strip(), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def read_payables(path):
    """Rows from the 'ready for deduction' CSV.

    The file leads with a title line and a blank line before the real header,
    and every data row carries a trailing empty field.
    """
    with open(path, encoding='utf-8-sig') as handle:
        lines = [line for line in handle.read().splitlines() if line.strip()]

    reader = csv.reader(lines)
    header = None
    rows = []
    for parts in reader:
        if not parts or not parts[0]:
            continue
        if header is None:
            if parts[0].strip() == 'Vendor code':
                header = parts
            continue
        rows.append({
            'vendor': parts[0].strip(),
            'txn': parts[1].strip(),
            'date': parse_date(parts[2]),
            'type': parts[3].strip(),
            'invoice': money(parts[4]),
            'deducted': money(parts[5]),
            'balance': money(parts[6]),
        })
    if header is None:
        raise SystemExit(f'No "Vendor code" header row found in {path}')
    return rows


def read_remittance(path):
    """(payments, lines) from the remittance workbook."""
    workbook = openpyxl.load_workbook(path, data_only=True)
    grid = list(workbook.active.iter_rows(values_only=True))

    pay_hdr = inv_hdr = None
    for index, row in enumerate(grid):
        if not row or not row[0]:
            continue
        first = str(row[0]).strip()
        if first == PAYMENT_HDR:
            if inv_hdr is None and row[1] == INVOICE_HDR_MARKER:
                inv_hdr = index
            elif pay_hdr is None:
                pay_hdr = index
    if pay_hdr is None or inv_hdr is None:
        raise SystemExit(f'Could not locate both section headers in {path}')

    payments = []
    for row in grid[pay_hdr + 1:inv_hdr]:
        if not row or not row[0]:
            continue
        number = str(row[0]).strip()
        # Two non-payment rows sit inside this span: a warning row that reuses
        # the 'Payment Number' label, and the 'Invoices' section title. Both
        # leave the amount column empty, which real payment rows never do.
        if number == PAYMENT_HDR or row[3] is None:
            continue
        payments.append({
            'number': number,
            'date': parse_date(row[1]),
            'currency': row[2],
            'amount': money(row[3]),
            'type': row[7],
            'status': row[8],
            'void_reason': row[9],
        })

    lines = []
    for row in grid[inv_hdr + 1:]:
        if not row or not row[0]:
            continue
        lines.append({
            'payment': str(row[0]).strip(),
            'invoice': str(row[1]).strip(),
            'date': parse_date(row[2]),
            'description': str(row[3] or ''),
            'amount': money(row[4]),
            'currency': row[5],
            'withholding': money(row[6]),
            'discount': money(row[7]),
            'paid': money(row[8]),
            'remaining': money(row[9]),
        })
    return payments, lines


def hidden_line_payments(path):
    """Payment numbers Amazon flags as having undisplayable remittance items."""
    grid = list(openpyxl.load_workbook(path, data_only=True).active.iter_rows(values_only=True))
    flagged = set()
    for row in grid:
        if not row or not row[0]:
            continue
        if str(row[0]).strip() != PAYMENT_HDR or not row[1]:
            continue
        # The warning row repeats the 'Payment Number' label but carries a
        # comma-separated list of payment numbers where a header label would
        # normally sit; that all-digits test is what separates the two.
        parts = [part.strip() for part in str(row[1]).split(',')]
        if all(part.isdigit() for part in parts):
            flagged.update(parts)
    return flagged


def build(payables, payments, lines, flagged):
    by_invoice = defaultdict(list)
    for line in lines:
        by_invoice[line['invoice']].append(line)
    payment_date = {p['number']: p['date'] for p in payments}

    matched, still_open = [], []
    for row in payables:
        hits = by_invoice.get(row['txn'])
        (matched if hits else still_open).append((row, hits or []))

    payable_txns = {row['txn'] for row in payables}
    orphan_deductions = [
        line for line in lines
        if line['amount'] is not None and line['amount'] < 0
        and line['invoice'] not in payable_txns
    ]

    applied = defaultdict(Decimal)
    for line in lines:
        if line['paid'] is not None:
            applied[line['payment']] += line['paid']

    tie_out = []
    for payment in payments:
        total = applied.get(payment['number'])
        # A payment with no line detail at all is short by its whole amount,
        # so the Difference column sums to the workbook-level gap either way.
        difference = payment['amount'] - (total or Decimal(0))
        tie_out.append({
            **payment,
            'line_total': total,
            'difference': difference,
            'flagged': payment['number'] in flagged,
        })

    sliced = []
    for invoice, hits in by_invoice.items():
        if len({hit['payment'] for hit in hits}) < 2:
            continue
        claim = hits[0]['amount']
        paid = sum(hit['paid'] for hit in hits if hit['paid'] is not None)
        discount = sum(hit['discount'] for hit in hits if hit['discount'] is not None)
        sliced.append({
            'invoice': invoice,
            'claim': claim,
            'slices': len(hits),
            'paid': paid,
            'discount': discount,
            'residual': (claim - discount - paid) if claim is not None else None,
            'payments': ', '.join(sorted({hit['payment'] for hit in hits})),
            'description': hits[0]['description'],
        })
    sliced.sort(key=lambda item: abs(item['claim'] or 0), reverse=True)

    return {
        'matched': matched,
        'still_open': still_open,
        'orphan_deductions': orphan_deductions,
        'tie_out': tie_out,
        'sliced': sliced,
        'payment_date': payment_date,
    }


HEAD_FILL = PatternFill('solid', fgColor='1F3A5F')
HEAD_FONT = Font(color='FFFFFF', bold=True)
THIN = Side(style='thin', color='D9D9D9')


def sheet(workbook, title, headers, rows, widths=None, money_cols=()):
    ws = workbook.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    for row in rows:
        ws.append(row)
    for index, width in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for column in money_cols:
        for cell in ws[get_column_letter(column)][1:]:
            cell.number_format = MONEY
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = Border(bottom=THIN)
    ws.freeze_panes = 'A2'
    if rows:
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{ws.max_row}'
    return ws


def write_workbook(result, payables, payments, lines, flagged, out_path,
                   payables_path, payments_path):
    total = lambda seq: sum(x for x in seq if x is not None)
    matched, still_open = result['matched'], result['still_open']
    payment_date = result['payment_date']

    matched_remitted = total(hit['paid'] for _, hits in matched for hit in hits)
    matched_balance = total(row['balance'] for row, _ in matched)
    dates = [p['date'] for p in payments if p['date']]
    window = f'{min(dates)} to {max(dates)}' if dates else 'unknown'

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    # ---- Summary -----------------------------------------------------------
    ws = workbook.create_sheet('Summary')
    ws.append(['Amazon payables vs. remittance detail'])
    ws['A1'].font = Font(size=14, bold=True)
    for label, value in [
        ('', ''),
        ('Payables file', payables_path),
        ('Remittance file', payments_path),
        ('Remittance window (payment dates)', window),
        ('', ''),
        ('OPEN PAYABLES AS EXPORTED', ''),
        ('Lines', len(payables)),
        ('Invoice amount', total(r['invoice'] for r in payables)),
        ('Already deducted per the payables file', total(r['deducted'] for r in payables)),
        ('Open balance', total(r['balance'] for r in payables)),
        ('', ''),
        ('MATCHED TO A REMITTANCE (tab 1)', ''),
        ('Lines', len(matched)),
        ('Invoice amount', total(r['invoice'] for r, _ in matched)),
        ('Amount the remittances actually applied', matched_remitted),
        ('Balance still shown as open -> overstatement', matched_balance),
        ('', ''),
        ('NO REMITTANCE HIT IN THE WINDOW (tab 2)', ''),
        ('Lines', len(still_open)),
        ('Invoice amount', total(r['invoice'] for r, _ in still_open)),
        ('Open balance', total(r['balance'] for r, _ in still_open)),
        ('', ''),
        ('REMITTANCE SIDE', ''),
        ('Payments', len(payments)),
        ('Payments total', total(p['amount'] for p in payments)),
        ('Remittance lines', len(lines)),
        ('Invoices settled (positive lines)',
         total(l['paid'] for l in lines if l['paid'] and l['paid'] > 0)),
        ('Deductions taken (negative lines)',
         total(l['paid'] for l in lines if l['paid'] and l['paid'] < 0)),
        ('Net of all lines', total(l['paid'] for l in lines)),
        ('Unexplained gap vs. payments total',
         total(p['amount'] for p in payments) - total(l['paid'] for l in lines)),
        ('Payments Amazon flagged as having hidden lines', len(flagged)),
        ('Payments whose lines tie out exactly',
         sum(1 for t in result['tie_out']
             if t['difference'] is not None and abs(t['difference']) < Decimal('0.005'))),
        ('Deductions taken that are not on the payables list (tab 3)',
         len(result['orphan_deductions'])),
    ]:
        ws.append([label, value])
    for cell in ws['B']:
        if isinstance(cell.value, Decimal):
            cell.number_format = MONEY
    for cell in ws['A']:
        if cell.value and str(cell.value).isupper():
            cell.font = Font(bold=True)
    ws.column_dimensions['A'].width = 52
    ws.column_dimensions['B'].width = 46

    # ---- 1. Already deducted ----------------------------------------------
    rows = []
    for row, hits in matched:
        for hit in hits:
            rows.append([
                row['txn'], row['date'], row['type'], row['invoice'],
                row['deducted'], row['balance'], hit['payment'],
                payment_date.get(hit['payment']), hit['paid'], hit['description'],
            ])
    rows.sort(key=lambda r: (r[0], str(r[7])))
    sheet(workbook, '1 Already deducted',
          ['Transaction number', 'Transaction date', 'Type', 'Invoice amount',
           'Deducted per payables', 'Balance shown open', 'Payment number',
           'Payment date', 'Amount applied', 'Remittance description'],
          rows, [24, 15, 15, 14, 16, 15, 15, 13, 14, 46], (4, 5, 6, 9))

    # ---- 2. Still open -----------------------------------------------------
    rows = sorted(
        ([row['txn'], row['date'], row['type'], row['invoice'],
          row['deducted'], row['balance']] for row, _ in still_open),
        key=lambda r: (r[2], r[5] or 0))
    sheet(workbook, '2 Still open',
          ['Transaction number', 'Transaction date', 'Type', 'Invoice amount',
           'Deducted', 'Open balance'],
          rows, [24, 15, 16, 14, 12, 14], (4, 5, 6))

    # ---- 3. Deducted but not on the payables list --------------------------
    rows = sorted(
        ([l['payment'], payment_date.get(l['payment']), l['invoice'], l['date'],
          l['amount'], l['paid'], l['description']]
         for l in result['orphan_deductions']),
        key=lambda r: (r[5] or 0))
    sheet(workbook, '3 Deducted not on list',
          ['Payment number', 'Payment date', 'Invoice number', 'Invoice date',
           'Claim amount', 'Amount applied', 'Description'],
          rows, [15, 13, 26, 13, 15, 15, 52], (5, 6))

    # ---- 4. Payment tie-out ------------------------------------------------
    rows = [[t['number'], t['date'], t['amount'], t['line_total'], t['difference'],
             'yes' if t['flagged'] else '',
             'no detail' if t['line_total'] is None
             else ('ties' if abs(t['difference']) < Decimal('0.005') else 'short'),
             t['type'], t['status']]
            for t in sorted(result['tie_out'], key=lambda t: (t['date'] or dt.date.min, t['number']))]
    sheet(workbook, '4 Payment tie-out',
          ['Payment number', 'Payment date', 'Payment amount', 'Sum of its lines',
           'Difference', 'Amazon flagged hidden lines', 'Result', 'Type', 'Status'],
          rows, [15, 13, 15, 16, 13, 14, 11, 8, 11], (3, 4, 5))

    # ---- 5. Multi-payment claims ------------------------------------------
    rows = [[s['invoice'], s['claim'], s['slices'], s['paid'], s['discount'],
             s['residual'],
             'ties' if s['residual'] is not None and abs(s['residual']) < Decimal('0.005')
             else 'CHECK',
             s['payments'], s['description']]
            for s in result['sliced']]
    sheet(workbook, '5 Multi-payment claims',
          ['Invoice number', 'Full claim', 'Slices', 'Total applied',
           'Terms discount', 'Residual', 'Result', 'Payments', 'Description'],
          rows, [34, 15, 8, 15, 14, 12, 9, 46, 46], (2, 4, 5, 6))

    workbook.save(out_path)
    return out_path


def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('payables', help='"Your payables - ready for deduction" CSV')
    parser.add_argument('payments', help='"Payments" XLSX from the remittance page')
    parser.add_argument('--out', default='amazon-payables-remittance-match.xlsx')
    args = parser.parse_args()

    payables = read_payables(args.payables)
    payments, lines = read_remittance(args.payments)
    flagged = hidden_line_payments(args.payments)
    result = build(payables, payments, lines, flagged)
    out = write_workbook(result, payables, payments, lines, flagged, args.out,
                         args.payables, args.payments)

    total = lambda seq: sum(x for x in seq if x is not None)
    print(f'payables lines        {len(payables):>6}   {total(r["invoice"] for r in payables):>14,.2f}')
    print(f'remittance payments   {len(payments):>6}   {total(p["amount"] for p in payments):>14,.2f}')
    print(f'remittance lines      {len(lines):>6}')
    print(f'  already deducted    {len(result["matched"]):>6}   '
          f'{total(r["balance"] for r, _ in result["matched"]):>14,.2f} still shown open')
    print(f'  no remittance hit   {len(result["still_open"]):>6}   '
          f'{total(r["balance"] for r, _ in result["still_open"]):>14,.2f}')
    print(f'  deducted, not listed{len(result["orphan_deductions"]):>6}')
    ties = sum(1 for t in result['tie_out']
               if t['difference'] is not None and abs(t['difference']) < Decimal('0.005'))
    print(f'  payments tying out  {ties:>6} / {len(payments)}')
    print(f'wrote {out}')


if __name__ == '__main__':
    main()
