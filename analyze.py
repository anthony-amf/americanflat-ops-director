"""Marketplace CPU analyzer.

Reads 3PL shipped orders + FedEx + Stamps.com invoices from a /uploads
directory and produces a CPU report for a chosen marketplace / carrier /
date window. Optionally compares to a pre-cutoff window and projects
annualized impact.

Designed as a read-only, ad-hoc analysis tool. Does not mutate any state.
"""
import argparse, csv, json, os, re, datetime as dt, sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import xlrd
except ImportError:
    xlrd = None

# --- Marketplace + carrier normalization (same as shipping-cost-report) ---

MKT = {
    "TARG":"Target", "TRGT":"Target", "TARGET":"Target",
    "MICHAELS":"Michaels", "MCHL":"Michaels",
    "SHOPIFY":"Shopify", "SHPFY":"Shopify",
    "MACY":"Macy's", "MACYS":"Macy's",
}

LTL_CARRIERS = {
    "WARD","FXFE","FXNL","FEDEX FREIGHT","ABFS","ABF FREIGHT",
    "RDWY","ROADWAY","YRC","YRCW","ODFL","OLD DOMINION",
    "SAIA","ESTES","RLCA","RL CARRIERS","TFRT","TFORCE",
    "PITD","PITT-OHIO","PITT OHIO",
    "GLOBAL E","GLOBAL-E","GLOBALE","GLOBAL_E",
    "9999",
}


def norm_mkt(code):
    if not code: return None
    s = str(code).strip().upper()
    if s.startswith("S-"): s = s[2:]
    return MKT.get(s)


def is_parcel(carrier):
    if not carrier: return True
    return str(carrier).strip().upper() not in LTL_CARRIERS


def clean_tr(tr):
    if not tr: return ""
    return re.sub(r"[^A-Z0-9]", "", str(tr).strip().upper())


def parse_date(s):
    if s is None: return None
    if isinstance(s, dt.datetime): return s.date()
    if isinstance(s, dt.date): return s
    s = str(s).strip()
    for fmt in ("%Y%m%d","%m/%d/%Y","%m/%d/%y","%Y-%m-%d","%m-%d-%Y","%Y-%m-%d %H:%M:%S"):
        try: return dt.datetime.strptime(s, fmt).date()
        except ValueError: pass
    return None


def week_monday(d):
    return d - dt.timedelta(days=d.weekday())


def week_label(m, today=None):
    e = m + dt.timedelta(days=6)
    if today and m <= today <= e:
        return f"W/O {m.month}/{m.day} thru {today.month}/{today.day}"
    return f"W/O {m.month}/{m.day} thru {e.month}/{e.day}"


# --- Loaders ---

def load_3pl_csv(path, orders, seen, mkt_filter):
    with open(path, encoding="utf-8-sig", errors="replace") as f:
        for row in csv.DictReader(f):
            ship = parse_date(row.get("Ship Date"))
            if not ship: continue
            tr = clean_tr(row.get("Bill of Lading"))
            mkt = norm_mkt(row.get("Batch#"))
            try: units = int(row.get("Units") or 0)
            except (ValueError, TypeError): units = 0
            carrier = (row.get("Carrier") or "").strip()
            if not (tr and mkt and units > 0): continue
            if mkt_filter and mkt != mkt_filter and mkt_filter != "All": continue
            if not is_parcel(carrier): continue
            k = (tr, ship)
            if k in seen: continue
            seen.add(k)
            if tr in orders:
                orders[tr]["units"] += units
            else:
                orders[tr] = {"mkt": mkt, "units": units, "ship": ship,
                              "carrier": carrier, "wh": "NJ/FON"}


def load_3pl_xlsx(path, orders, seen, mkt_filter):
    if openpyxl is None: return
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    hdr = next(it)
    idx = {h: i for i, h in enumerate(hdr)}
    tr_col, dv_col = idx.get("Tracking Number"), idx.get("Division")
    if tr_col is None or dv_col is None: return
    qty_col = idx.get("Shipped Quantity")
    sd_col = idx.get("Actual Ship Date")
    car_col = idx.get("Carrier")
    sku_col = idx.get("SKU/STYLE")
    line_col = idx.get("Line No.")
    for r in it:
        if not r: continue
        ship = parse_date(r[sd_col]) if sd_col is not None else None
        if not ship: continue
        tr = clean_tr(r[tr_col])
        mkt = norm_mkt(r[dv_col])
        try: qty = int(r[qty_col] or 0)
        except (ValueError, TypeError): qty = 0
        if not (tr and mkt and qty > 0): continue
        if mkt_filter and mkt != mkt_filter and mkt_filter != "All": continue
        carrier = r[car_col] if car_col is not None else ""
        if not is_parcel(carrier): continue
        sku = r[sku_col] if sku_col is not None else None
        line = r[line_col] if line_col is not None else None
        k = (tr, ship, "SC", sku, line, qty)
        if k in seen: continue
        seen.add(k)
        if tr in orders:
            orders[tr]["units"] += qty
        else:
            orders[tr] = {"mkt": mkt, "units": qty, "ship": ship,
                          "carrier": str(carrier), "wh": "SC"}


def load_fedex(path, fedex, seen):
    low = path.lower()
    try:
        if low.endswith(".csv"):
            with open(path, encoding="utf-8-sig", errors="replace") as f:
                for row in csv.DictReader(f):
                    tr = clean_tr(row.get("Express or Ground Tracking ID"))
                    if not tr: continue
                    try: net = float(row.get("Net Charge Amount") or 0)
                    except (ValueError, TypeError): net = 0
                    ship = parse_date(row.get("Shipment Date"))
                    k = (tr, round(net, 2), ship)
                    if k in seen: continue
                    seen.add(k); fedex[tr] += net
        elif low.endswith(".xlsx") and openpyxl:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            hdr = next(it)
            idx = {h: i for i, h in enumerate(hdr)}
            tc = idx.get("Express or Ground Tracking ID")
            nc = idx.get("Net Charge Amount")
            sc = idx.get("Shipment Date")
            if tc is None or nc is None: return
            for row in it:
                if not row: continue
                tr = clean_tr(row[tc])
                if not tr: continue
                try: net = float(row[nc] or 0)
                except (ValueError, TypeError): net = 0
                ship = parse_date(row[sc]) if sc is not None else None
                k = (tr, round(net, 2), ship)
                if k in seen: continue
                seen.add(k); fedex[tr] += net
        elif low.endswith(".xls") and xlrd:
            wb = xlrd.open_workbook(path, on_demand=True)
            sh = wb.sheet_by_index(0)
            hdr = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
            idx = {h: i for i, h in enumerate(hdr)}
            tc = idx.get("Express or Ground Tracking ID")
            nc = idx.get("Net Charge Amount")
            sd = idx.get("Shipment Date")
            if tc is None or nc is None: return
            for r in range(1, sh.nrows):
                tr = clean_tr(sh.cell_value(r, tc))
                if not tr: continue
                try: net = float(sh.cell_value(r, nc) or 0)
                except (ValueError, TypeError): net = 0
                ship = parse_date(sh.cell_value(r, sd)) if sd is not None else None
                k = (tr, round(net, 2), ship)
                if k in seen: continue
                seen.add(k); fedex[tr] += net
    except Exception as e:
        print(f"WARN: failed to read {os.path.basename(path)}: {e}", file=sys.stderr)


def load_stamps(path, stamps):
    with open(path, encoding="utf-8-sig", errors="replace") as f:
        for row in csv.DictReader(f):
            tr = clean_tr(row.get("Tracking #") or row.get("Tracking Number"))
            if not tr: continue
            try: amt = float(row.get("Amount Paid") or row.get("Adjusted Amount") or 0)
            except (ValueError, TypeError): amt = 0
            stamps[tr] += amt


# --- File classifier ---

def classify(path):
    up = os.path.basename(path).upper()
    if "PRINTHISTORY" in up or "STAMPS" in up: return "stamps"
    if "FEDEX" in up: return "fedex"
    if up.endswith(".CSV"):
        try:
            with open(path, encoding="utf-8-sig") as f:
                cols = {c.strip() for c in next(csv.reader(f))}
            if {"Bill of Lading","Batch#","Ship Date","Units"}.issubset(cols):
                return "nj_fon"
            if "Tracking #" in cols: return "stamps"
            if "Express or Ground Tracking ID" in cols: return "fedex"
        except Exception:
            pass
    if up.endswith(".XLSX") and openpyxl:
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            cols = {(h or "").strip() for h in next(wb.active.iter_rows(values_only=True))}
            if {"Tracking Number","Division","Shipped Quantity"}.issubset(cols):
                return "sc"
            if {"Express or Ground Tracking ID","Net Charge Amount"}.issubset(cols):
                return "fedex"
        except Exception:
            pass
    if up.endswith(".XLS") and xlrd:
        try:
            wb = xlrd.open_workbook(path, on_demand=True)
            sh = wb.sheet_by_index(0)
            cols = {str(sh.cell_value(0,c)).strip() for c in range(sh.ncols)}
            if {"Express or Ground Tracking ID","Net Charge Amount"}.issubset(cols):
                return "fedex"
        except Exception:
            pass
    return None


# --- Analysis ---

def analyze_window(orders, fedex, stamps, since, until, carrier_filter):
    """Return per-week dict with totals + the filtered (matched-by-carrier) subset."""
    out = defaultdict(lambda: {
        "total_un": 0, "total_trk": 0,
        "matched_un": 0, "matched_trk": 0, "matched_cost": 0.0,
        "fedex_un": 0, "fedex_cost": 0.0,
        "stamps_un": 0, "stamps_cost": 0.0,
    })
    for tr, o in orders.items():
        if not (since <= o["ship"] <= until): continue
        wk = week_monday(o["ship"])
        out[wk]["total_un"] += o["units"]
        out[wk]["total_trk"] += 1
        in_fedex = tr in fedex and fedex[tr] > 0
        in_stamps = tr in stamps
        stamps_cost = max(stamps[tr], 0) if in_stamps else 0  # refund-zero
        if in_fedex:
            out[wk]["fedex_un"] += o["units"]
            out[wk]["fedex_cost"] += fedex[tr]
        elif in_stamps:
            out[wk]["stamps_un"] += o["units"]
            out[wk]["stamps_cost"] += stamps_cost
        # Apply carrier filter to "matched"
        cf = carrier_filter.lower() if carrier_filter else "all"
        if cf == "fedex":
            if in_fedex:
                out[wk]["matched_un"] += o["units"]
                out[wk]["matched_trk"] += 1
                out[wk]["matched_cost"] += fedex[tr]
        elif cf == "stamps":
            if in_stamps and not in_fedex:
                out[wk]["matched_un"] += o["units"]
                out[wk]["matched_trk"] += 1
                out[wk]["matched_cost"] += stamps_cost
        else:
            if in_fedex:
                out[wk]["matched_un"] += o["units"]
                out[wk]["matched_trk"] += 1
                out[wk]["matched_cost"] += fedex[tr]
            elif in_stamps:
                out[wk]["matched_un"] += o["units"]
                out[wk]["matched_trk"] += 1
                out[wk]["matched_cost"] += stamps_cost
    return out


def cpu(c, u): return c/u if u else 0


def render_report(args, weekly, weekly_pre):
    lines = []
    mk = args.marketplace
    car = args.carrier or "All"
    lines.append(f"# {mk} CPU Analysis — {car} carrier")
    lines.append("")
    lines.append(f"**Window:** {args.since} → {args.until or 'today'}")
    lines.append("")

    # Weekly breakdown
    lines.append("## Weekly breakdown")
    lines.append("")
    lines.append("| Week | Total units | Matched units | %Coverage | Cost | CPU |")
    lines.append("|---|---:|---:|---:|---:|---:|")
    total_u = matched_u = 0; matched_c = 0.0; matched_trk = 0
    for wk in sorted(weekly):
        v = weekly[wk]
        pct = 100*v["matched_un"]/v["total_un"] if v["total_un"] else 0
        c = cpu(v["matched_cost"], v["matched_un"])
        lines.append(f"| W/O {wk.month}/{wk.day} | {v['total_un']} | {v['matched_un']} | {pct:.0f}% | ${v['matched_cost']:,.2f} | ${c:.2f} |")
        total_u += v["total_un"]; matched_u += v["matched_un"]
        matched_c += v["matched_cost"]; matched_trk += v["matched_trk"]
    lines.append("")

    # Summary
    summary_cpu = cpu(matched_c, matched_u)
    cov = 100*matched_u/total_u if total_u else 0
    lines.append("## Window summary")
    lines.append("")
    lines.append(f"- **{matched_trk:,} trackings / {matched_u:,} units invoiced on {car}**")
    lines.append(f"- **Total {car} cost:** ${matched_c:,.2f}")
    lines.append(f"- **{car} CPU:** **${summary_cpu:.2f}**")
    lines.append(f"- Coverage: {matched_u:,} of {total_u:,} total {mk} units ({cov:.1f}%)")
    if cov < 95:
        lines.append(f"- *Note: coverage <95% — recent weeks likely have invoice lag. Wait 1-2 weeks for full picture.*")
    lines.append("")

    # Pre/post compare
    if weekly_pre is not None or args.baseline_cpu is not None:
        lines.append("## Pre/post comparison")
        lines.append("")
        if args.baseline_cpu is not None:
            pre_cpu = float(args.baseline_cpu)
            pre_note = "(historical reference rate — provided via --baseline-cpu)"
            pre_units = pre_cost = None
        else:
            pre_units = sum(v["matched_un"] for v in weekly_pre.values())
            pre_cost  = sum(v["matched_cost"] for v in weekly_pre.values())
            pre_cpu   = cpu(pre_cost, pre_units)
            pre_note = ""
        delta = summary_cpu - pre_cpu
        mult = summary_cpu / pre_cpu if pre_cpu else 0
        lines.append(f"| Period | CPU | Units | Cost |")
        lines.append(f"|---|---:|---:|---:|")
        if pre_units is not None:
            lines.append(f"| Pre (before {args.compare_before or args.since}) | ${pre_cpu:.2f} | {pre_units:,} | ${pre_cost:,.2f} |")
        else:
            lines.append(f"| Pre (reference) {pre_note} | ${pre_cpu:.2f} | — | — |")
        lines.append(f"| Post (since {args.since}) | ${summary_cpu:.2f} | {matched_u:,} | ${matched_c:,.2f} |")
        lines.append(f"| **Δ CPU** | **${delta:+.2f}** ({mult:.2f}×) | | |")
        lines.append("")

        # Annual projection
        weeks_in_window = max(1, len([w for w in weekly if weekly[w]["matched_un"] > 0]))
        avg_weekly_units = matched_u / weeks_in_window if weeks_in_window else 0
        annual_units = avg_weekly_units * 52
        annual_extra = delta * annual_units
        if annual_extra != 0:
            sign = "more" if annual_extra > 0 else "less"
            lines.append("## Annualized impact (flat-volume projection)")
            lines.append("")
            lines.append(f"- Average weekly volume in window: **{avg_weekly_units:.0f} units**")
            lines.append(f"- × 52 weeks = {annual_units:,.0f} units/year")
            lines.append(f"- × delta CPU (${delta:+.2f}) = **${abs(annual_extra):,.0f}/year {sign} than pre rate**")
            lines.append("")
            lines.append("*Caveat: assumes flat volume at the current window's pace. If the routing/volume mix is still changing, this projection lags reality.*")
    return "\n".join(lines) + "\n"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--uploads", required=True)
    ap.add_argument("--marketplace", required=True,
                    help="Target | Michaels | Shopify | Macy's | All")
    ap.add_argument("--carrier", default="All",
                    help="FedEx | Stamps | All (default All)")
    ap.add_argument("--since", required=True, help="ISO date, inclusive")
    ap.add_argument("--until", default=None, help="ISO date, inclusive; default today")
    ap.add_argument("--compare-before", default=None,
                    help="ISO date — everything before this becomes the 'pre' comparison")
    ap.add_argument("--baseline-cpu", default=None, type=float,
                    help="Override pre CPU with an explicit reference rate")
    ap.add_argument("--out", default=None, help="Markdown output path")
    args = ap.parse_args()

    since = dt.date.fromisoformat(args.since)
    until = dt.date.fromisoformat(args.until) if args.until else dt.date.today()
    compare_before = dt.date.fromisoformat(args.compare_before) if args.compare_before else None

    if args.marketplace != "All" and args.marketplace not in {"Target","Michaels","Shopify","Macy's"}:
        sys.exit(f"Bad marketplace: {args.marketplace}")

    # Load files
    orders = {}; seen = set()
    fedex = defaultdict(float); fx_seen = set()
    stamps = defaultdict(float)
    nj_fon_paths = []
    for fn in os.listdir(args.uploads):
        full = os.path.join(args.uploads, fn)
        if not os.path.isfile(full): continue
        cls = classify(full)
        if cls == "fedex":
            load_fedex(full, fedex, fx_seen)
        elif cls == "stamps":
            load_stamps(full, stamps)
        elif cls == "sc":
            load_3pl_xlsx(full, orders, seen, args.marketplace)
        elif cls == "nj_fon":
            nj_fon_paths.append(full)
    for p in nj_fon_paths:
        load_3pl_csv(p, orders, seen, args.marketplace)

    # Refund-zero rule
    for tr in list(stamps.keys()):
        if stamps[tr] <= 0: stamps[tr] = 0.0

    print(f"Loaded: {len(orders)} {args.marketplace} 3PL orders | "
          f"{len(fedex)} FedEx trackings | {len(stamps)} Stamps trackings",
          file=sys.stderr)

    # Build windows
    weekly = analyze_window(orders, fedex, stamps, since, until, args.carrier)

    weekly_pre = None
    if compare_before:
        # Pre window: everything before compare_before
        pre_end = compare_before - dt.timedelta(days=1)
        pre_start = dt.date(2025, 1, 1)
        weekly_pre = analyze_window(orders, fedex, stamps, pre_start, pre_end, args.carrier)

    report = render_report(args, weekly, weekly_pre)
    if args.out:
        with open(args.out, "w") as f:
            f.write(report)
        print(f"Wrote {args.out}", file=sys.stderr)
    print(report)


if __name__ == "__main__":
    main()
