#!/usr/bin/env python3
"""
Parse supporting Excel files from Yusen invoices.

Extracts order numbers, quantities, and service types from the Excel files
that accompany Small Parcel/LTL invoices.

Usage:
    python scripts/parse_invoice_excel.py path/to/invoice.xlsx 751996
    python scripts/parse_invoice_excel.py path/to/invoice.xlsx 751996 --output orders.json
"""

import sys
import json
import argparse
from pathlib import Path
from typing import Any

import openpyxl


def detect_format(wb) -> str:
    """Detect Excel file format: 'yusen' or 'taylored_services'."""
    sheetnames = wb.sheetnames

    # Yusen format: has "Small Parcel" or "SML PRCL" sheet
    if any(s in sheetnames for s in ["Small Parcel", "SML PRCL", "LTL"]):
        return "yusen"

    # Taylored Services format: has "Sheet1", "Sheet2" with TSI PO# header
    if "Sheet1" in sheetnames:
        ws = wb["Sheet1"]
        first_cell = ws["A1"].value
        if first_cell and "TSI" in str(first_cell).upper():
            return "taylored_services"

    return "yusen"  # Default


def parse_taylored_services_excel(file_path: str, invoice_number: str, warehouse: str = "FONTANA") -> dict[str, Any]:
    """
    Parse Taylored Services supporting Excel file.

    Expected structure:
    - Sheet1: E-commerce/small parcel orders starting at row 14, order in column A
    - Sheet2: LTL BOLs starting at row 8, BOL in column B

    Returns:
        {
            "invoice_number": "752319",
            "warehouse": "FONTANA",
            "line_items": [
                {"order_number": "102003276483843", "service_type": "Small Parcel", "quantity": 1},
                ...
            ]
        }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    line_items = []

    # Parse Sheet1: E-commerce/Small Parcel orders (starting row 14)
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=14, max_row=ws.max_row, min_col=1, max_col=10), start=1):
            order_num = row[0].value  # Column A
            if order_num and str(order_num).strip() and str(order_num).strip() != "Order":
                line_items.append({
                    "line_item_id": row_idx,
                    "order_number": str(order_num).strip(),
                    "quantity": 1,
                    "service_type": "Small Parcel",
                    "warehouse_location": warehouse,
                })

    # Parse Sheet2: LTL BOLs (starting row 8, BOL in column B)
    if "Sheet2" in wb.sheetnames:
        ws = wb["Sheet2"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=10), start=1000):
            bol_num = row[1].value if len(row) > 1 else None  # Column B
            if bol_num and str(bol_num).strip():
                line_items.append({
                    "line_item_id": row_idx + 1000,
                    "order_number": str(bol_num).strip(),
                    "quantity": 1,
                    "service_type": "LTL",
                    "warehouse_location": warehouse,
                    "notes": "BOL number",
                })

    return {
        "invoice_number": invoice_number,
        "warehouse_location": warehouse,
        "total_line_items": len(line_items),
        "line_items": line_items,
    }


def parse_yusen_excel(file_path: str, invoice_number: str, warehouse: str = "NEW JERSEY") -> dict[str, Any]:
    """
    Parse Yusen supporting Excel file.

    Expected structure:
    - "Small Parcel" sheet with order numbers in column A
    - "LTL" sheet (if present) with SSCC/BOL numbers

    Returns:
        {
            "invoice_number": "751996",
            "warehouse": "NEW JERSEY",
            "line_items": [
                {"order_number": "1Z...", "service_type": "Small Parcel", "quantity": 1},
                ...
            ]
        }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    line_items = []

    # Parse Small Parcel sheet
    if "Small Parcel" in wb.sheetnames or "SML PRCL" in wb.sheetnames:
        sheet_name = "Small Parcel" if "Small Parcel" in wb.sheetnames else "SML PRCL"
        ws = wb[sheet_name]

        # Extract order numbers from column A (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5), start=1):
            order_num = row[0].value
            quantity = row[1].value if len(row) > 1 else None
            notes = row[4].value if len(row) > 4 else None

            if order_num:  # Skip empty rows
                line_items.append({
                    "line_item_id": row_idx,
                    "order_number": str(order_num).strip(),
                    "quantity": int(quantity) if quantity else 1,
                    "service_type": "Small Parcel",
                    "warehouse_location": warehouse,
                    "notes": notes,
                })

    # Parse LTL sheet
    if "LTL" in wb.sheetnames:
        ws = wb["LTL"]

        # Extract BOL/SSCC numbers from column A
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5), start=1000):
            bol_num = row[0].value
            quantity = row[1].value if len(row) > 1 else None

            if bol_num:
                line_items.append({
                    "line_item_id": row_idx,
                    "order_number": str(bol_num).strip(),
                    "quantity": int(quantity) if quantity else 1,
                    "service_type": "LTL",
                    "warehouse_location": warehouse,
                    "notes": "BOL/SSCC number",
                })

    return {
        "invoice_number": invoice_number,
        "warehouse_location": warehouse,
        "total_line_items": len(line_items),
        "line_items": line_items,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Parse invoice supporting Excel files (Yusen or Taylored Services)",
        epilog="""
Examples:
  python scripts/parse_invoice_excel.py samples/751996.xlsx 751996
  python scripts/parse_invoice_excel.py samples/752319.xlsx 752319 --warehouse "FONTANA"
  python scripts/parse_invoice_excel.py samples/751542.xlsx 751542 --output orders.json
        """
    )
    parser.add_argument("excel_file", help="Path to Excel file")
    parser.add_argument("invoice_number", help="Invoice number (e.g., 751996)")
    parser.add_argument("--warehouse", help="Warehouse location (auto-detected if not specified)")
    parser.add_argument("--output", help="Output JSON file (default: print to stdout)")

    args = parser.parse_args()

    # Verify file exists
    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        print(f"Error: File not found: {args.excel_file}", file=sys.stderr)
        sys.exit(1)

    try:
        # Detect format
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        fmt = detect_format(wb)

        # Set warehouse default based on format
        warehouse = args.warehouse or ("FONTANA" if fmt == "taylored_services" else "NEW JERSEY")

        # Parse with appropriate parser
        if fmt == "taylored_services":
            result = parse_taylored_services_excel(str(excel_path), args.invoice_number, warehouse)
        else:
            result = parse_yusen_excel(str(excel_path), args.invoice_number, warehouse)

        # Output
        output_text = json.dumps(result, indent=2)

        if args.output:
            with open(args.output, "w") as f:
                f.write(output_text)
            print(f"✓ Parsed {result['total_line_items']} line items ({fmt}) to {args.output}")
        else:
            print(output_text)

    except Exception as e:
        print(f"Error parsing Excel: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
