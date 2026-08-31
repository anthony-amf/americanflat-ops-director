#!/usr/bin/env python3
"""Build the Marketplace Shipments portal from BigQuery.

Pulls order/shipment records for the four marketplaces Americanflat ships
directly (Target, Macy's, Michaels, Shopify) out of the marketplace feeds that
already land in BigQuery, normalizes them onto one shipment-shaped row, and
writes a self-contained HTML page that mirrors the Yusen invoice portal.

Sources (one per marketplace):
  Target    acenda.ship_advice_raw + acenda.fulfillment_raw   (Target Plus US)
  Macy's    macys.orders_clean                                (Mirakl feed)
  Michaels  shipstation.orders_clean  storeName 'AMF Michaels'
  Shopify   shipstation.orders_clean  storeName 'Shopify'

Known gap: ShipStation's *shipment* feed (shipstation.shipments_raw) stopped
loading in Oct 2023, so Michaels and Shopify rows carry no ship date and no
tracking number — only the order, the customer, and the requested carrier.
Target and Macy's are complete. See MARKETPLACE-SHIPMENTS.md.

Usage:
    python3 refresh_marketplace_shipments.py                 # 180 days -> HTML
    python3 refresh_marketplace_shipments.py --days 365
    python3 refresh_marketplace_shipments.py --ndjson out.ndjson   # for a BQ load
"""

import argparse
import csv
import datetime as dt
import json
import os
import re
import shutil
import subprocess
import sys
import urllib.request

PROJECT = "americanflat"
BQ = "https://bigquery.googleapis.com/bigquery/v2"
EPOCH = dt.date(2020, 1, 1)

MARKETPLACES = ["Target", "Macy's", "Michaels", "Shopify"]


# --------------------------------------------------------------------------
# BigQuery
# --------------------------------------------------------------------------
def access_token(mode):
    """Bearer token, or None when the agent proxy injects credentials for us.

    In a Claude cloud session bigquery.googleapis.com is an injected host, so we
    send no Authorization header. On the Mac there is no proxy, so we borrow the
    gcloud ADC token.
    """
    if mode == "proxy":
        return None
    if mode == "gcloud" or (mode == "auto" and not os.environ.get("HTTPS_PROXY")):
        if not shutil.which("gcloud"):
            sys.exit("gcloud not found — run with --auth proxy inside a cloud session.")
        return subprocess.run(["gcloud", "auth", "print-access-token"],
                              capture_output=True, text=True, check=True).stdout.strip()
    return None


def _post(url, payload, token):
    data = json.dumps(payload).encode() if payload is not None else None
    req = urllib.request.Request(url, data=data, method="POST" if data else "GET")
    req.add_header("Content-Type", "application/json")
    if token:
        req.add_header("Authorization", "Bearer " + token)
    with urllib.request.urlopen(req, timeout=300) as resp:
        return json.load(resp)


def query(sql, token, page_size=20000):
    """Run a query and return a list of dicts, following pagination."""
    body = {"query": sql, "useLegacySql": False, "timeoutMs": 120000,
            "maxResults": page_size}
    res = _post(f"{BQ}/projects/{PROJECT}/queries", body, token)
    job = res["jobReference"]
    fields = [f["name"] for f in res.get("schema", {}).get("fields", [])]
    rows, page = [], res

    def collect(p):
        for r in p.get("rows", []):
            rows.append({k: c["v"] for k, c in zip(fields, r["f"])})

    collect(page)
    while not page.get("jobComplete", True) or page.get("pageToken"):
        params = f"?maxResults={page_size}"
        if page.get("pageToken"):
            params += "&pageToken=" + page["pageToken"]
        loc = job.get("location", "US")
        page = _post(f"{BQ}/projects/{PROJECT}/queries/{job['jobId']}{params}&location={loc}",
                     None, token)
        if not fields:
            fields = [f["name"] for f in page.get("schema", {}).get("fields", [])]
        collect(page)
    return rows


# --------------------------------------------------------------------------
# The unified shipment query
# --------------------------------------------------------------------------
SQL = r"""
WITH
-- ---- Target: Acenda ship advices (item grain) + fulfillments (tracking) ----
tgt_orders AS (
  SELECT
    shipAdviceId,
    ANY_VALUE(purchaseOrder)                                   AS order_number,
    ANY_VALUE(orderId)                                         AS order_ref,
    ANY_VALUE(status)                                          AS raw_status,
    SAFE_CAST(SUBSTR(ANY_VALUE(createdAt), 1, 10) AS DATE)     AS order_date,
    NULLIF(TRIM(CONCAT(COALESCE(ANY_VALUE(shipToFirstName), ''), ' ',
                       COALESCE(ANY_VALUE(shipToLastName), ''))), '')  AS customer,
    ANY_VALUE(shipToCity)                                      AS city,
    ANY_VALUE(shipToState)                                     AS state
  FROM `americanflat.acenda.ship_advice_raw`
  GROUP BY shipAdviceId
),
tgt_ship AS (
  -- One order can have several fulfillments. Take the earliest as *the* shipment
  -- and read its carrier and tracking off that same row: pulling each field with
  -- ANY_VALUE independently pairs a UPS tracking number with a FedEx label.
  SELECT
    shipAdviceId,
    SAFE_CAST(SUBSTR(first.dateShipped, 1, 10) AS DATE)        AS ship_date,
    first.carrier                                              AS carrier,
    first.trackingNumber                                       AS tracking
  FROM (
    SELECT
      shipAdviceId,
      ARRAY_AGG(STRUCT(dateShipped, carrier, trackingNumber)
                ORDER BY dateShipped, trackingNumber LIMIT 1)[OFFSET(0)] AS first
    FROM `americanflat.acenda.fulfillment_raw`
    WHERE status != 'canceled'
    GROUP BY shipAdviceId
  )
),
tgt AS (
  SELECT 'Target' AS marketplace, o.order_number, o.order_ref, o.customer, o.city, o.state,
         o.order_date, s.ship_date, o.raw_status, s.carrier, s.tracking,
         i.sku, i.productName AS product, i.quantity AS qty,
         i.unitPrice AS unit_price, i.totalItemPrice AS line_total
  FROM `americanflat.acenda.ship_advice_raw` i
  JOIN tgt_orders o USING (shipAdviceId)
  LEFT JOIN tgt_ship s USING (shipAdviceId)
),

-- ---- Macy's: Mirakl order feed (line grain) ----
mac_orders AS (
  SELECT
    orderId                                                    AS order_ref,
    COALESCE(ANY_VALUE(commercialId), orderId)                 AS order_number,
    NULLIF(TRIM(CONCAT(COALESCE(ANY_VALUE(shipToFirstName), ''), ' ',
                       COALESCE(ANY_VALUE(shipToLastName), ''))), '')  AS customer,
    ANY_VALUE(shipToCity)                                      AS city,
    ANY_VALUE(shipToState)                                     AS state,
    ANY_VALUE(orderDay)                                        AS order_date,
    DATE(MIN(shippedDate))                                     AS ship_date,
    ANY_VALUE(shippingCompany)                                 AS carrier,
    NULLIF(ANY_VALUE(shippingTracking), '')                    AS tracking,
    ANY_VALUE(orderState)                                      AS raw_status
  FROM `americanflat.macys.orders_clean`
  GROUP BY orderId
),
mac AS (
  SELECT "Macy's" AS marketplace, o.order_number, o.order_ref, o.customer, o.city, o.state,
         o.order_date, o.ship_date, o.raw_status, o.carrier, o.tracking,
         -- productShopSku is our own SKU; productSku is Macy's internal id.
         COALESCE(NULLIF(l.productShopSku, ''), l.offerSku, l.productSku) AS sku,
         l.productTitle AS product, l.quantity AS qty,
         l.linePrice AS unit_price, l.lineTotalPrice AS line_total
  FROM `americanflat.macys.orders_clean` l
  JOIN mac_orders o ON o.order_ref = l.orderId
),

-- ---- Michaels + Shopify: ShipStation order feed (no shipment feed since 2023) ----
ss_orders AS (
  SELECT
    orderId,
    CASE WHEN ANY_VALUE(storeName) LIKE '%Michaels%' THEN 'Michaels' ELSE 'Shopify' END AS marketplace,
    ANY_VALUE(orderNumber)                                     AS order_number,
    NULLIF(TRIM(COALESCE(ANY_VALUE(shipToName), '')), '')      AS customer,
    ANY_VALUE(shipToCity)                                      AS city,
    ANY_VALUE(shipToState)                                     AS state,
    ANY_VALUE(orderDay)                                        AS order_date,
    ANY_VALUE(COALESCE(NULLIF(carrierCode, ''), requestedShippingService))  AS carrier,
    ANY_VALUE(orderStatus)                                     AS raw_status
  FROM `americanflat.shipstation.orders_clean`
  WHERE storeName IN ('AMF Michaels', 'Manual Michaels Orders',
                      'Shopify', 'Manual Shopify Orders')
  GROUP BY orderId
),
ss AS (
  SELECT o.marketplace, o.order_number, CAST(o.orderId AS STRING) AS order_ref,
         o.customer, o.city, o.state, o.order_date,
         CAST(NULL AS DATE) AS ship_date, o.raw_status, o.carrier,
         CAST(NULL AS STRING) AS tracking,
         l.sku, l.itemName AS product, l.quantity AS qty,
         l.unitPrice AS unit_price, l.lineTotal AS line_total
  FROM `americanflat.shipstation.orders_clean` l
  JOIN ss_orders o ON o.orderId = l.orderId
  WHERE l.storeName IN ('AMF Michaels', 'Manual Michaels Orders',
                        'Shopify', 'Manual Shopify Orders')
)

-- One row per order line. The three CTEs project the same columns in the same
-- order, so a positional UNION ALL is safe; the builder groups them into orders.
SELECT * FROM (
  SELECT * FROM tgt
  UNION ALL SELECT * FROM mac
  UNION ALL SELECT * FROM ss
)
WHERE COALESCE(ship_date, order_date) >= DATE_SUB(CURRENT_DATE(), INTERVAL @DAYS DAY)
ORDER BY COALESCE(ship_date, order_date) DESC
"""


LEDGER_SQL = r"""
SELECT t.marketplace, t.order_number, t.order_ref, t.customer, t.city, t.state,
       t.order_date, t.ship_date, t.carrier, t.tracking, t.status AS raw_status,
       it.sku, it.product, it.qty, it.unit_price, it.line_total
FROM `americanflat.marketplaces.marketplace_shipments` t
LEFT JOIN UNNEST(t.items) it
WHERE COALESCE(t.ship_date, t.order_date) >= DATE_SUB(CURRENT_DATE(), INTERVAL @DAYS DAY)
ORDER BY COALESCE(t.ship_date, t.order_date) DESC
"""


def build_sql(days, source="feeds"):
    """`feeds` reads the marketplace tables directly; `ledger` reads the durable
    table the same rows get merged into (see sql/marketplace_shipments_setup.sql),
    which keeps customer names Target has since redacted at the source."""
    sql = LEDGER_SQL if source == "ledger" else SQL
    return sql.replace("@DAYS", str(int(days)))


# --------------------------------------------------------------------------
# Normalization
# --------------------------------------------------------------------------
CARRIERS = {
    "fedx": "FedEx", "fedex": "FedEx", "fedex_walleted": "FedEx",
    "ups": "UPS", "ups_walleted": "UPS",
    "usps": "USPS", "stamps_com": "USPS", "stamps.com": "USPS",
    "dhl": "DHL", "dhl_express_worldwide": "DHL",
}

STATUS = {
    "shipped": "Shipped", "closed": "Shipped", "shipping": "Shipping",
    "received": "Open", "pending": "Open", "awaiting_shipment": "Open",
    "awaiting_payment": "Open", "on_hold": "On hold",
    "canceled": "Cancelled", "cancelled": "Cancelled",
}

TRACK_URL = {
    "UPS": "https://www.ups.com/track?tracknum=",
    "FedEx": "https://www.fedex.com/fedextrack/?trknbr=",
    "USPS": "https://tools.usps.com/go/TrackConfirmAction?tLabels=",
}


def norm_carrier(raw):
    if not raw:
        return ""
    key = raw.strip().lower()
    if key in CARRIERS:
        return CARRIERS[key]
    for token, name in (("fedex", "FedEx"), ("fedx", "FedEx"), ("usps", "USPS"),
                        ("stamps", "USPS"), ("dhl", "DHL"), ("ups", "UPS")):
        if token in key:
            return name
    # ShipStation stores the storefront rate name ("Free Shipping - $65+") when no
    # carrier was assigned. That is a price, not a carrier — leave it blank rather
    # than let it crowd the carrier panel.
    return ""


def norm_status(raw, ship_date):
    if not raw:
        return "Shipped" if ship_date else "Open"
    s = STATUS.get(raw.strip().lower())
    if s:
        # A Mirakl order sits in RECEIVED even after it ships; the ship date is
        # the fact that matters, so let it win over the marketplace's own state.
        if s == "Open" and ship_date:
            return "Shipped"
        return s
    return raw.strip().title()


# --------------------------------------------------------------------------
# Parcel charges — what the carrier billed us, keyed by tracking number
# --------------------------------------------------------------------------
# No feed carries this: acenda's fulfillment cost is 0.00 on every row and
# ShipStation's shipment feed (which had shipmentCost) died in Oct 2023. The
# real number is on the FedEx and Stamps.com invoices, which is why the weekly
# shipping-cost report matches them to orders by tracking number. Same join here.
#
# Four layouts are accepted: the two consolidated Drive sheets, and the raw
# exports the weekly download drops in the uploads folder.
COST_LAYOUTS = [
    # (tracking column, amount column, carrier column or fixed carrier, date column)
    ("Tracking Number", "Net Charge", "FedEx", "Invoice Date"),               # AMF FedEx Invoices sheet
    ("Tracking Number", "Amount Paid", ("col", "Carrier"), "Ship Date"),      # AMF Stamps.com Invoices sheet
    ("Express or Ground Tracking ID", "Net Charge Amount", "FedEx", "Invoice Date"),  # FedEx Billing Online
    ("Tracking #", "Amount Paid", ("col", "Carrier"), "Ship Date"),           # Stamps.com print history
]


def norm_tracking(value):
    """Keep letters and digits only.

    Stamps.com exports USPS numbers Excel-escaped as ="9434650105798022300341"
    while UPS numbers come through bare, so a laxer strip silently matched every
    UPS shipment and no USPS one.
    """
    return re.sub(r"[^A-Za-z0-9]", "", str(value or "")).upper()


def parse_amount(value):
    try:
        return float(str(value).replace("$", "").replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _cost_rows(path):
    """Yield (tracking, amount, carrier, charge date) from one invoice export."""
    if path.lower().endswith((".xlsx", ".xls")):
        rows = _read_xlsx(path)
    else:
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
            rows = list(csv.DictReader(fh))
    if not rows:
        return
    header = set(rows[0].keys())
    for tcol, acol, carrier, dcol in COST_LAYOUTS:
        if tcol in header and acol in header:
            break
    else:
        print("  skipped %s (no tracking/amount columns)" % os.path.basename(path),
              file=sys.stderr)
        return
    for r in rows:
        # A refunded label costs us nothing. Stamps keeps the row with its original
        # Amount Paid, so counting it would bill a shipment we were credited for.
        # (Amount Paid already includes any re-rate: quoted + adjusted = paid.)
        if (str(r.get("Refund Status", "")).strip().lower() == "approved"
                or str(r.get("Shipment Status", "")).strip().lower() == "refunded"):
            continue
        tracking = norm_tracking(r.get(tcol))
        amount = parse_amount(r.get(acol))
        # The consolidated sheets are several exports stacked, so a repeated
        # header row shows up as data. A non-numeric amount is that, not a charge.
        if not tracking or amount is None:
            continue
        name = r.get(carrier[1]) if isinstance(carrier, tuple) else carrier
        yield tracking, amount, norm_carrier(name) or "Other", str(r.get(dcol, "") or "")


def _read_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl is needed to read %s — pass the CSV export instead." % path)
    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows)]
    return [dict(zip(header, r)) for r in rows]


def load_costs(paths):
    """(tracking -> {"cost", "source"}, list of the distinct charge lines).

    Two rules, and both matter:

    1. Charges are SUMMED per tracking number. FedEx bills a shipment again when
       it re-rates one (a dimensional-weight correction, an address surcharge),
       and what we pay is the sum of those lines.

    2. A charge is counted ONCE per (tracking, date, amount). The consolidated
       Drive sheets are weekly exports stacked on top of each other and the
       downloads overlap, so the same invoice line is physically present two or
       three times — 4,266 of the FedEx sheet's 7,963 rows are repeats. Summing
       them raw put Target's FedEx cost per unit at $35.92 against a known ~$13.
    """
    files = []
    for p in paths:
        if os.path.isdir(p):
            files += [os.path.join(p, f) for f in sorted(os.listdir(p))
                      if f.lower().endswith((".csv", ".xlsx", ".xls"))]
        else:
            files.append(p)
    seen, costs, lines, repeats = set(), {}, [], 0
    for path in files:
        n = 0
        for tracking, amount, carrier, date in _cost_rows(path):
            key = (tracking, date, amount)
            if key in seen:
                repeats += 1
                continue
            seen.add(key)
            entry = costs.setdefault(tracking, {"cost": 0.0, "source": carrier})
            entry["cost"] = round(entry["cost"] + amount, 2)
            lines.append({"tracking": tracking, "amount": round(amount, 2),
                          "carrier": carrier, "charge_date": iso_date(date)})
            n += 1
        print("  %-46s %6d charges" % (os.path.basename(path)[:46], n), file=sys.stderr)
    if repeats:
        print("  %6d repeated invoice lines ignored" % repeats, file=sys.stderr)
    return costs, lines


def iso_date(value):
    """Invoice dates arrive as 20260202 or 01/05/2026 depending on the export."""
    v = str(value or "").strip()
    for fmt in ("%Y%m%d", "%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return dt.datetime.strptime(v, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


# --------------------------------------------------------------------------
# 3PL shipped-order reports — the ship date and tracking ShipStation stopped sending
# --------------------------------------------------------------------------
# Michaels and Shopify orders reach BigQuery from ShipStation, whose shipment
# feed died in Oct 2023, so they arrive with no ship date and no tracking number
# — and with no tracking there is nothing for a carrier invoice to match, which
# is why they were showing no cost either. The warehouses do report all three:
# the weekly NJ / Fontana / South Carolina shipped-order reports are keyed by
# order number. Fill the gap from those and the cost join then works normally.
THREEPL_LAYOUTS = [
    # (order column, tracking column, ship date column, carrier column)
    ("Order Number", "Tracking Number", "Ship Date", "Carrier"),        # consolidated Drive sheet
    ("Order", "Bill of Lading", "Ship Date", "Carrier"),                # NJ / Fontana export
    ("Order No.", "Tracking Number", "Actual Ship Date", "Carrier"),    # South Carolina export
]


def order_key(value):
    """A Michaels order is THP6600107706404869-2 in ShipStation and plain
    6600107706404869 at the warehouse; Shopify is 23901 in both. Reduce to the
    digits both sides agree on."""
    v = re.sub(r"\s", "", str(value or "")).upper()
    v = re.sub(r"^THP", "", v)
    v = re.sub(r"-\d+$", "", v)
    return v


def load_3pl(paths):
    """order key -> {"ship_date", "tracking", "carrier"} from warehouse reports."""
    files = []
    for p in paths:
        if os.path.isdir(p):
            files += [os.path.join(p, f) for f in sorted(os.listdir(p))
                      if f.lower().endswith((".csv", ".xlsx", ".xls"))]
        else:
            files.append(p)
    index = {}
    for path in files:
        if path.lower().endswith((".xlsx", ".xls")):
            rows = _read_xlsx(path)
        else:
            with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
                rows = list(csv.DictReader(fh))
        if not rows:
            continue
        header = set(rows[0].keys())
        for ocol, tcol, dcol, ccol in THREEPL_LAYOUTS:
            if ocol in header and tcol in header:
                break
        else:
            print("  skipped %s (no order/tracking columns)" % os.path.basename(path),
                  file=sys.stderr)
            continue
        n = 0
        for r in rows:
            key = order_key(r.get(ocol))
            tracking = norm_tracking(r.get(tcol))
            if not key or not tracking:
                continue
            ship = iso_date(r.get(dcol))
            prior = index.get(key)
            # A split order has a row per carton. Keep the first one that shipped,
            # to match how the Target feed reports a multi-fulfillment order.
            if prior and prior["ship_date"] and (not ship or prior["ship_date"] <= ship):
                continue
            index[key] = {"ship_date": ship, "tracking": tracking,
                          "carrier": norm_carrier(r.get(ccol))}
            n += 1
        print("  %-46s %6d orders" % (os.path.basename(path)[:46], n), file=sys.stderr)
    return index


def attach_3pl(rows, index):
    """Fill in only what the marketplace feed left blank — never overwrite it."""
    filled = 0
    for r in rows:
        if r["tracking"] and r["ship_date"]:
            continue
        hit = index.get(order_key(r["order_number"])) or index.get(order_key(r["order_ref"]))
        if not hit:
            continue
        if not r["tracking"] and hit["tracking"]:
            r["tracking"] = hit["tracking"]
            r["shipped_by"] = "3PL report"
            filled += 1
        if not r["ship_date"] and hit["ship_date"]:
            r["ship_date"] = hit["ship_date"]
            r["status"] = norm_status(None, hit["ship_date"])
        if not r["carrier"] and hit["carrier"]:
            r["carrier"] = hit["carrier"]
    return filled


# --------------------------------------------------------------------------
# The order-level shipping cost report — costs already matched to orders
# --------------------------------------------------------------------------
# The weekly shipping cost pipeline emits a reconciled per-order roll-up
# (all_orders_shipping_costs_<date>.md): order number, marketplace, ship date,
# carrier and the summed cost of every package on the order. It reaches orders
# this builder cannot price on its own — the ones with no tracking number in
# BigQuery at all — so it is joined by order number as a second cost source.
#
# Precedence: a tracking-level match wins where we have one. Both sources agree
# to the cent on 85% of the orders they share; where they differ it is usually a
# Stamps adjustment that posted after the report was built (its Quoted Amount
# plus a later $7.50 adjustment is our Amount Paid), so the fresher per-shipment
# figure is the one to keep.
ORDER_COST_COLUMNS = ("Order", "Ship Date", "Carrier", "Cost")


def load_order_costs(paths):
    """order key -> {"cost", "carrier", "ship_date"} from the shipping cost report."""
    index = {}
    for path in paths:
        n = 0
        header = None
        for line in open(path, encoding="utf-8", errors="replace"):
            if not line.startswith("|"):
                continue
            cells = [c.strip() for c in line.strip().strip("|").split("|")]
            if header is None:
                if "Order" in cells and "Cost" in cells:
                    header = cells
                continue
            # The row of dashes under a markdown header, and the summary tables
            # above the detail one, are not orders.
            if len(cells) != len(header) or set("".join(cells)) <= set("-: "):
                continue
            row = dict(zip(header, cells))
            key = order_key(row.get("Order"))
            cost = parse_amount(str(row.get("Cost", "")).replace("$", ""))
            if not key or cost is None:
                continue
            index[key] = {"cost": round(cost, 2),
                          "carrier": norm_carrier(row.get("Carrier")),
                          "ship_date": iso_date(row.get("Ship Date"))}
            n += 1
        print("  %-46s %6d orders" % (os.path.basename(path)[:46], n), file=sys.stderr)
    return index


def attach_order_costs(rows, index):
    """Price the orders the tracking join could not reach, and fill their dates."""
    priced = 0
    for r in rows:
        hit = index.get(order_key(r["order_number"])) or index.get(order_key(r["order_ref"]))
        if not hit:
            continue
        if not r["ship_date"] and hit["ship_date"]:
            r["ship_date"] = hit["ship_date"]
            r["status"] = norm_status(None, hit["ship_date"])
        if not r["carrier"] and hit["carrier"]:
            r["carrier"] = hit["carrier"]
        if r.get("ship_cost") is None:
            r["ship_cost"] = hit["cost"]
            # Say which join produced the figure: one is per shipment, the other
            # is the whole order, and on a split order those differ legitimately.
            r["cost_source"] = (hit["carrier"] or "Carrier") + " (order match)"
            priced += 1
    return priced


COST_TABLE_SQL = r"""
SELECT tracking, SUM(amount) AS cost, ANY_VALUE(carrier) AS source
FROM `%s`
GROUP BY tracking
"""


def attach_costs(rows, costs):
    """Put the carrier charge on each order, by its tracking number."""
    for r in rows:
        hit = costs.get(norm_tracking(r["tracking"])) if r["tracking"] else None
        r["ship_cost"] = round(hit["cost"], 2) if hit else None
        r["cost_source"] = hit["source"] if hit else ""
    return rows


def title_case(name):
    """3PL and marketplace feeds mix ALL CAPS with Mixed Case; even them out."""
    if not name:
        return ""
    n = name.strip()
    if n.isupper() or n.islower():
        n = " ".join(w.capitalize() for w in n.split())
    return n


def normalize(rows):
    """Collapse line-grain rows into one order each, carrying its line items."""
    orders = {}
    for r in rows:
        key = (r["marketplace"], r.get("order_ref") or r.get("order_number") or "")
        o = orders.get(key)
        if o is None:
            ship, order = r.get("ship_date"), r.get("order_date")
            # A redacted marketplace name is not a name — Target Plus swaps consumer
            # PII for "Customer Name" ~45 days after the order, so don't pretend.
            cust = title_case(r.get("customer") or "")
            if cust.lower() in ("customer name", "name customer", "customer"):
                cust = ""
            o = orders[key] = {
                "marketplace": r["marketplace"],
                "order_number": (r.get("order_number") or "").strip(),
                "order_ref": (r.get("order_ref") or "").strip(),
                "customer": cust,
                "city": title_case(r.get("city") or ""),
                "state": (r.get("state") or "").strip().upper()[:2],
                "order_date": order or "",
                "ship_date": ship or "",
                "carrier": norm_carrier(r.get("carrier")),
                "tracking": (r.get("tracking") or "").strip(),
                "status": norm_status(r.get("raw_status"), ship),
                "items": [],
            }
        sku = (r.get("sku") or "").strip()
        product = (r.get("product") or "").strip()
        qty = int(float(r.get("qty") or 0))
        if not (sku or product or qty):
            continue
        o["items"].append({
            "sku": sku, "product": product, "qty": qty,
            "unit_price": round(float(r.get("unit_price") or 0), 2),
            "line_total": round(float(r.get("line_total") or 0), 2),
        })

    out = []
    for o in orders.values():
        # ShipStation carries discounts as line items with a name but no SKU
        # ("WELCOME10", qty 1, -$8.59). They belong in the order's value, which is
        # what the customer paid, but counting one as a shipped unit is wrong.
        o["units"] = sum(i["qty"] for i in o["items"] if i["sku"])
        o["skus"] = len({i["sku"] for i in o["items"] if i["sku"]})
        # Order value is the sum of its own lines, so the column and the expanded
        # detail can never disagree.
        o["value"] = round(sum(i["line_total"] for i in o["items"]), 2)
        out.append(o)
    return out


# --------------------------------------------------------------------------
# Compact encoding for the page
# --------------------------------------------------------------------------
def day_num(iso):
    if not iso:
        return -1
    return (dt.date.fromisoformat(iso) - EPOCH).days


def encode(rows):
    """Dictionary-encode everything repetitive: 40k orders and 50k line items go
    into the page as arrays of small integers plus one products table."""
    mkts, carriers, statuses, states, products, sources = [], [], [], [], [], [""]
    pidx = {}

    def idx(pool, val):
        if val not in pool:
            pool.append(val)
        return pool.index(val)

    def product_idx(sku, name):
        key = (sku, name)
        if key not in pidx:
            pidx[key] = len(products)
            products.append([sku, name])
        return pidx[key]

    data = []
    for r in rows:
        items = [[product_idx(i["sku"], i["product"]), i["qty"],
                  i["unit_price"], i["line_total"]] for i in r["items"]]
        data.append([
            day_num(r["ship_date"]), day_num(r["order_date"]),
            idx(mkts, r["marketplace"]), r["order_number"], r["customer"],
            r["city"], idx(states, r["state"]), r["units"], r["skus"],
            idx(carriers, r["carrier"]), r["tracking"], idx(statuses, r["status"]),
            r["value"], items,
            # null, not 0: no invoice for this shipment is a different fact from a
            # shipment that cost nothing, and the page has to show them differently.
            r.get("ship_cost"), idx(sources, r.get("cost_source") or ""),
        ])
    return {"rows": data, "mkt": mkts, "carrier": carriers, "status": statuses,
            "state": states, "products": products, "source": sources,
            "epoch": EPOCH.isoformat()}


def kpi_block(rows, days, filled_3pl=0):
    dates = sorted(d for d in (r["ship_date"] or r["order_date"] for r in rows) if d)
    # The window the invoices actually cover, so the page can say so rather than
    # leaving a half-empty column unexplained.
    cost_dates = sorted(r["ship_date"] for r in rows
                        if r.get("ship_cost") is not None and r["ship_date"])
    by_m = {}
    for r in rows:
        m = by_m.setdefault(r["marketplace"],
                            {"n": 0, "units": 0, "tracked": 0, "shipdate": 0,
                             "value": 0.0, "costed": 0, "cost": 0.0})
        m["n"] += 1
        m["units"] += r["units"]
        m["value"] = round(m["value"] + r["value"], 2)
        m["tracked"] += 1 if r["tracking"] else 0
        m["shipdate"] += 1 if r["ship_date"] else 0
        if r.get("ship_cost") is not None:
            m["costed"] += 1
            m["cost"] = round(m["cost"] + r["ship_cost"], 2)
    return {
        "n": len(rows), "days": days,
        "date_lo": dates[0] if dates else "", "date_hi": dates[-1] if dates else "",
        "built": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "by_marketplace": by_m,
        "filled_3pl": filled_3pl,
        "cost_lo": cost_dates[0] if cost_dates else "",
        "cost_hi": cost_dates[-1] if cost_dates else "",
    }


# --------------------------------------------------------------------------
# The page.  Mirrors the Yusen invoice portal: same tokens, same KPI cards,
# same bar panels, same sticky toolbar + sortable table.
# --------------------------------------------------------------------------
TEMPLATE = r"""<title>Marketplace Shipments</title>
<style>
  :root {
    --paper: #F6F8F9;
    --panel: #FFFFFF;
    --ink: #1B2129;
    --muted: #5C6873;
    --line: #E3E7EB;
    --track: #EDF0F2;
    --thead: #F1F4F6;
    --accent: #14627E;
    --accent-soft: rgba(20, 98, 126, 0.10);
    --shadow: 0 1px 2px rgba(20, 30, 40, 0.05), 0 1px 3px rgba(20, 30, 40, 0.08);
    --m-Target: #CC2229;
    --m-Macys: #7C3AED;
    --m-Michaels: #1565C0;
    --m-Shopify: #15803D;
    --dash: #C4CBD1;
    --ok: #15803D;
    --warn: #B45309;
  }
  @media (prefers-color-scheme: dark) {
    :root:not([data-theme="light"]) {
      --paper: #14181D;
      --panel: #1B2129;
      --ink: #E8ECEF;
      --muted: #98A4AE;
      --line: #2A313A;
      --track: #232A32;
      --thead: #20262E;
      --accent: #4FA3C2;
      --accent-soft: rgba(79, 163, 194, 0.14);
      --shadow: 0 1px 2px rgba(0, 0, 0, 0.35);
      --m-Target: #F0888D;
      --m-Macys: #B08AF8;
      --m-Michaels: #6BAEE8;
      --m-Shopify: #5BC57E;
      --dash: #46505A;
      --ok: #5BC57E;
      --warn: #F5A94B;
    }
  }
  :root[data-theme="dark"] {
    --paper: #14181D;
    --panel: #1B2129;
    --ink: #E8ECEF;
    --muted: #98A4AE;
    --line: #2A313A;
    --track: #232A32;
    --thead: #20262E;
    --accent: #4FA3C2;
    --accent-soft: rgba(79, 163, 194, 0.14);
    --shadow: 0 1px 2px rgba(0, 0, 0, 0.35);
    --m-Target: #F0888D;
    --m-Macys: #B08AF8;
    --m-Michaels: #6BAEE8;
    --m-Shopify: #5BC57E;
    --dash: #46505A;
    --ok: #5BC57E;
    --warn: #F5A94B;
  }

  /* The viewer paints its own ground behind the page, so claim it from a token
     rather than inheriting the host's theme. */
  body { background: var(--paper); color: var(--ink); }

  .mp-root {
    background: var(--paper); color: var(--ink);
    font: 14px/1.5 -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
    -webkit-font-smoothing: antialiased;
    min-height: 100vh;
  }
  .mp-wrap { max-width: 1800px; margin: 0 auto; padding: 28px 20px 64px; }

  .mp-head { display: flex; align-items: baseline; justify-content: space-between; flex-wrap: wrap; gap: 8px; margin-bottom: 4px; }
  .mp-head h1 { font-size: 25px; font-weight: 700; letter-spacing: -0.02em; margin: 0; }
  .mp-sub { color: var(--muted); font-size: 13px; }
  .mp-sub code { background: var(--accent-soft); color: var(--accent); padding: 1px 6px; border-radius: 5px; font-size: 12px; }

  .mp-kpis { display: grid; grid-template-columns: repeat(auto-fit, minmax(170px, 1fr)); gap: 14px; margin: 22px 0 26px; }
  .kpi { background: var(--panel); border: 1px solid var(--line); border-radius: 10px; padding: 15px 17px; box-shadow: var(--shadow); }
  .kpi .label { color: var(--muted); font-size: 11.5px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em; }
  .kpi .value { font-size: 23px; font-weight: 700; letter-spacing: -0.02em; margin-top: 6px; font-variant-numeric: tabular-nums; }
  .kpi .value small { font-size: 13px; font-weight: 500; color: var(--muted); letter-spacing: 0; }

  .mp-panels { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 26px; }
  @media (max-width: 860px) { .mp-panels { grid-template-columns: 1fr; } }
  .panel { background: var(--panel); border: 1px solid var(--line); border-radius: 10px; padding: 17px 19px; box-shadow: var(--shadow); }
  .panel h2 { font-size: 13.5px; font-weight: 650; margin: 0 0 2px; }
  .panel .hint { color: var(--muted); font-size: 11.5px; margin: 0 0 12px; }
  .bar-row { display: grid; grid-template-columns: 128px 1fr auto; align-items: center; gap: 10px; margin: 8px 0; }
  .bar-row .name { font-size: 12.5px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
  .bar-track { background: var(--track); border-radius: 5px; height: 17px; overflow: hidden; }
  .bar-fill { height: 100%; border-radius: 5px; }
  .bar-row .amt { font-variant-numeric: tabular-nums; font-size: 12.5px; color: var(--muted); white-space: nowrap; }
  @media (prefers-reduced-motion: no-preference) { .bar-fill { transition: width 0.25s ease; } }

  /* Search + filters ride at the top of the viewport; the column headers park
     directly under them (--toolbar-h is measured in JS because the bar wraps). */
  .mp-toolbar { display: flex; gap: 10px; flex-wrap: wrap; align-items: center;
    position: sticky; top: 0; z-index: 4; background: var(--paper);
    padding: 10px 0; margin-bottom: 2px; box-shadow: 0 1px 0 var(--line); }
  .mp-toolbar.nostick { position: static; box-shadow: none; }
  .mp-toolbar input[type="search"], .mp-toolbar select {
    font: inherit; padding: 8px 11px; border: 1px solid var(--line); border-radius: 8px;
    background: var(--panel); color: var(--ink);
  }
  .mp-toolbar input[type="search"] { flex: 1; min-width: 220px; }
  .mp-toolbar input[type="search"]:focus-visible, .mp-toolbar select:focus-visible,
  .mp-table thead th:focus-visible, a.link:focus-visible {
    outline: 2px solid var(--accent); outline-offset: 1px;
  }
  .mp-count { color: var(--muted); font-size: 13px; margin-left: auto; font-variant-numeric: tabular-nums; }

  .mp-tablewrap { background: var(--panel); border: 1px solid var(--line); border-radius: 10px; box-shadow: var(--shadow); overflow-x: auto; }
  @media (min-width: 1320px) { .mp-tablewrap { overflow-x: visible; } }
  .mp-table { border-collapse: collapse; width: 100%; min-width: 860px; }
  .mp-table thead th {
    position: sticky; top: var(--toolbar-h, 56px); background: var(--thead); z-index: 1;
    text-align: left; font-size: 11.5px; font-weight: 650; color: var(--muted);
    text-transform: uppercase; letter-spacing: 0.04em; padding: 10px 10px;
    border-bottom: 1px solid var(--line); cursor: pointer; user-select: none; white-space: nowrap;
  }
  .mp-table thead th.num { text-align: right; }
  .mp-table thead th .arrow { opacity: 0.4; font-size: 10px; margin-left: 3px; }
  .mp-table thead th.sorted .arrow { opacity: 1; color: var(--accent); }
  .mp-table tbody td { padding: 9px 10px; border-bottom: 1px solid var(--line); vertical-align: top; }
  .mp-table tbody tr:hover { background: var(--thead); }
  .mp-table td.num { text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }
  .mp-table td.ord { font-variant-numeric: tabular-nums; font-weight: 600; white-space: nowrap; }
  .mp-table td.date { white-space: nowrap; }
  .mp-table td.cust { max-width: 220px; overflow-wrap: anywhere; }
  .mp-table td.dest { color: var(--muted); white-space: nowrap; }

  .expbtn { background: none; border: 0; padding: 0; font: inherit; font-weight: 600; cursor: pointer;
    color: var(--ink); display: inline-flex; align-items: center; gap: 6px;
    font-variant-numeric: tabular-nums; white-space: nowrap; }
  .expbtn:hover { color: var(--accent); }
  .expbtn:focus-visible { outline: 2px solid var(--accent); outline-offset: 2px; border-radius: 6px; }
  .ecaret { color: var(--muted); font-size: 9px; transition: transform .12s ease; }
  @media (prefers-reduced-motion: reduce) { .ecaret { transition: none; } }
  .expbtn[aria-expanded="true"] .ecaret { transform: rotate(90deg); }
  tr.linedetail > td { background: var(--thead); padding: 12px 16px 14px; }
  .lines { border-collapse: collapse; width: 100%; max-width: 900px; }
  .lines th { text-align: left; font-size: 10.5px; font-weight: 700; text-transform: uppercase;
    letter-spacing: .05em; color: var(--muted); padding: 0 10px 5px 0; white-space: nowrap; }
  .lines th.num, .lines td.num { text-align: right; }
  .lines td { padding: 4px 10px 4px 0; border-top: 1px solid var(--line); vertical-align: top;
    font-size: 12.5px; }
  .lines td.num { font-variant-numeric: tabular-nums; white-space: nowrap; }
  .lines td.sku { font-weight: 600; font-variant-numeric: tabular-nums; white-space: nowrap; }
  .lines td.prod { color: var(--muted); }
  .lines tfoot td { border-top: 1px solid var(--line); font-weight: 700; padding-top: 6px; }
  .lines .vnone { color: var(--muted); font-style: italic; }
  .linecost { margin: 9px 0 0; font-size: 12.5px; color: var(--muted); font-variant-numeric: tabular-nums; }
  .linecost b { color: var(--ink); }

  .chip { display: inline-block; padding: 2px 9px; border-radius: 999px; font-size: 11.5px; font-weight: 600; white-space: nowrap; }
  a.link { color: var(--accent); text-decoration: none; font-weight: 600; white-space: nowrap; font-variant-numeric: tabular-nums; }
  a.link:hover { text-decoration: underline; }
  .dash { color: var(--dash); }
  .mp-more { display: flex; justify-content: center; padding: 14px 0 0; }
  /* display:flex outranks the hidden attribute, so say it again. */
  .mp-more[hidden] { display: none; }
  .morebtn { font: inherit; font-size: 13px; font-weight: 600; cursor: pointer; color: var(--accent);
    background: var(--accent-soft); border: 1px solid var(--line); border-radius: 8px; padding: 8px 18px; }
  .morebtn:hover { border-color: var(--accent); }
  .morebtn:focus-visible { outline: 2px solid var(--accent); outline-offset: 1px; }
  .mp-foot { color: var(--muted); font-size: 12px; margin-top: 18px; text-align: center; }
  .mp-foot p { margin: 4px 0; }
  .mp-note { background: var(--panel); border: 1px solid var(--line); border-left: 3px solid var(--warn);
    border-radius: 8px; padding: 11px 15px; margin: 0 0 18px; color: var(--muted); font-size: 12.5px; }
  .mp-note b { color: var(--ink); }
</style>

<div class="mp-root"><div class="mp-wrap">
  <header class="mp-head">
    <div>
      <h1>Marketplace Shipments</h1>
      <div class="mp-sub">Target &middot; Macy's &middot; Michaels &middot; Shopify &mdash; live from <code>americanflat</code> BigQuery</div>
    </div>
    <div class="mp-sub" id="meta"></div>
  </header>

  <section class="mp-kpis" id="kpis"></section>

  <section class="mp-panels">
    <div class="panel"><h2>Shipments by marketplace</h2><p class="hint" id="hint-mkt"></p><div id="chart-mkt"></div></div>
    <div class="panel"><h2>Shipments by carrier</h2><p class="hint" id="hint-car"></p><div id="chart-car"></div></div>
  </section>

  <p class="mp-note" id="gapnote"></p>

  <div class="mp-toolbar">
    <input type="search" id="q" placeholder="Search order #, customer name, tracking, SKU, city&#8230;" autocomplete="off" aria-label="Search shipments">
    <select id="f-mkt" aria-label="Filter by marketplace"><option value="">All marketplaces</option></select>
    <select id="f-month" aria-label="Filter by month"><option value="">All months</option></select>
    <select id="f-car" aria-label="Filter by carrier"><option value="">All carriers</option></select>
    <select id="f-status" aria-label="Filter by status"><option value="">All statuses</option></select>
    <select id="f-track" aria-label="Filter by tracking"><option value="">Tracking: any</option><option value="yes">Has tracking</option><option value="no">No tracking</option></select>
    <span class="mp-count" id="count"></span>
  </div>

  <div class="mp-tablewrap">
    <table class="mp-table">
      <thead><tr>
        <th data-k="ship" tabindex="0">Ship date<span class="arrow">&#9660;</span></th>
        <th data-k="order" tabindex="0">Ordered<span class="arrow">&#9660;</span></th>
        <th data-k="mkt" tabindex="0">Marketplace<span class="arrow">&#9660;</span></th>
        <th data-k="num" tabindex="0">Order #<span class="arrow">&#9660;</span></th>
        <th data-k="cust" tabindex="0">Customer<span class="arrow">&#9660;</span></th>
        <th data-k="dest" tabindex="0">Destination<span class="arrow">&#9660;</span></th>
        <th data-k="units" class="num" tabindex="0">Units<span class="arrow">&#9660;</span></th>
        <th data-k="value" class="num" tabindex="0">Order value<span class="arrow">&#9660;</span></th>
        <th data-k="cost" class="num" tabindex="0">Ship cost<span class="arrow">&#9660;</span></th>
        <th data-k="car" tabindex="0">Carrier<span class="arrow">&#9660;</span></th>
        <th data-k="trk" tabindex="0">Tracking<span class="arrow">&#9660;</span></th>
        <th data-k="st" tabindex="0">Status<span class="arrow">&#9660;</span></th>
      </tr></thead>
      <tbody id="rows"></tbody>
    </table>
  </div>

  <div class="mp-more" id="morewrap" hidden>
    <button class="morebtn" type="button" id="more"></button>
  </div>

  <footer class="mp-foot" id="foot"></footer>
</div></div>

<script>
const RAW = __DATA__;
const KPI = __KPI__;

const EPOCH = Date.parse(RAW.epoch + "T00:00:00Z");
const DAY = 86400000;
const MON = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"];
const esc = s => String(s == null ? "" : s).replace(/[&<>"]/g, c => ({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;"}[c]));
const num = n => n.toLocaleString("en-US");

// Rows arrive as dictionary-encoded arrays (40k+ of them), so unpack once into
// objects the filter/sort can read by name without re-indexing on every pass.
const C = {SHIP:0, ORDER:1, MKT:2, NUM:3, CUST:4, CITY:5, STATE:6, UNITS:7, SKUS:8,
           CAR:9, TRK:10, ST:11, VALUE:12, ITEMS:13, COST:14, CSRC:15};
const P = {SKU:0, NAME:1};   // products table
const L = {P:0, QTY:1, UNIT:2, TOTAL:3};  // one line item
const iso = d => d < 0 ? "" : new Date(EPOCH + d * DAY).toISOString().slice(0, 10);
const DATA = RAW.rows.map((r, i) => {
  const shipIso = iso(r[C.SHIP]), orderIso = iso(r[C.ORDER]);
  const state = RAW.state[r[C.STATE]] || "";
  const city = r[C.CITY] || "";
  return {
    i,
    ship: r[C.SHIP], order: r[C.ORDER], shipIso, orderIso,
    mkt: RAW.mkt[r[C.MKT]], num: r[C.NUM] || "", cust: r[C.CUST] || "",
    dest: city + (city && state ? ", " : "") + state,
    units: r[C.UNITS], skus: r[C.SKUS],
    car: RAW.carrier[r[C.CAR]] || "", trk: r[C.TRK] || "", st: RAW.status[r[C.ST]],
    value: r[C.VALUE], items: r[C.ITEMS] || [],
    // null means no invoice found for this tracking number, which is not $0.
    cost: r[C.COST] == null ? null : r[C.COST],
    csrc: RAW.source[r[C.CSRC]] || "",
    // One lowercase haystack per row, built once: search stays instant at 40k rows.
    // SKUs are in it (people look up "who bought MW0808DWOOD"); product titles are
    // not — they are long, near-duplicate, and would triple the page's memory.
    hay: ((r[C.NUM] || "") + " " + (r[C.CUST] || "") + " " + (r[C.TRK] || "") + " " +
          city + " " + state + " " + RAW.mkt[r[C.MKT]] + " " +
          (r[C.ITEMS] || []).map(l => RAW.products[l[L.P]][P.SKU]).join(" ")).toLowerCase(),
  };
});

const money = n => "$" + n.toLocaleString("en-US", {minimumFractionDigits: 2, maximumFractionDigits: 2});
const money0 = n => "$" + Math.round(n).toLocaleString("en-US");

const fmtDate = d => { if (!d) return ""; const [y, m, dd] = d.split("-"); return MON[+m - 1] + " " + dd + ", " + y.slice(2); };
const fmtMonth = ym => { const [y, m] = ym.split("-"); return MON[+m - 1] + " " + y; };
const mktVar = m => "--m-" + m.replace(/[^A-Za-z]/g, "");

const TRACK_URL = {
  UPS: "https://www.ups.com/track?tracknum=",
  FedEx: "https://www.fedex.com/fedextrack/?trknbr=",
  USPS: "https://tools.usps.com/go/TrackConfirmAction?tLabels=",
};

function carrierOf(r) {
  const t = (r.trk || "").replace(/\s/g, "").toUpperCase();
  if (/^1Z/.test(t)) return "UPS";
  if (/^(94|93|92|95|82)\d{18,}$/.test(t)) return "USPS";
  if (/^\d{12}$|^\d{15}$|^\d{20}$/.test(t)) return "FedEx";
  return r.car;
}

document.getElementById("meta").textContent =
  num(KPI.n) + " shipments \u00b7 " + fmtDate(KPI.date_lo) + " \u2192 " + fmtDate(KPI.date_hi);

// ---- KPI cards + bar panels, recomputed for the current filter ----
function kpiCards(rows) {
  const shipped = rows.filter(r => r.st === "Shipped").length;
  const tracked = rows.filter(r => r.trk).length;
  const units = rows.reduce((a, r) => a + r.units, 0);
  // Days-to-ship only means anything on rows that carry both dates.
  const spans = rows.filter(r => r.ship >= 0 && r.order >= 0).map(r => r.ship - r.order);
  const avgSpan = spans.length ? (spans.reduce((a, b) => a + b, 0) / spans.length) : null;
  const value = rows.reduce((a, r) => a + r.value, 0);
  const priced = rows.filter(r => r.cost != null);
  const cost = priced.reduce((a, r) => a + r.cost, 0);
  const pricedUnits = priced.reduce((a, r) => a + r.units, 0);
  const cards = [
    {label: "Shipments", value: num(rows.length), sub: rows.length === DATA.length ? "" : "of " + num(DATA.length)},
    {label: "Units", value: num(units)},
    {label: "Order value", value: money0(value),
     sub: rows.length ? money(value / rows.length) + " avg" : ""},
    {label: "Shipping cost", value: priced.length ? money0(cost) : "\u2014",
     sub: priced.length ? num(priced.length) + " of " + num(rows.length) + " priced" : "no invoices loaded"},
    // Not the weekly report's blended CPU: this is only the shipments an invoice
    // has been matched to, and it excludes nothing the way that report does.
    {label: "Cost per unit", value: pricedUnits ? money(cost / pricedUnits) : "\u2014",
     sub: pricedUnits ? "on " + num(pricedUnits) + " priced units" : ""},
    {label: "Shipped", value: num(shipped), sub: "of " + num(rows.length)},
    {label: "With tracking", value: num(tracked), sub: "of " + num(rows.length)},
    {label: "Avg days to ship", value: avgSpan == null ? "\u2014" : avgSpan.toFixed(1),
     sub: spans.length ? "on " + num(spans.length) + " rows" : "no ship dates"},
  ];
  document.getElementById("kpis").innerHTML = cards.map(k =>
    '<div class="kpi"><div class="label">' + k.label + '</div>' +
    '<div class="value">' + k.value + (k.sub ? ' <small>' + k.sub + '</small>' : "") + '</div></div>'
  ).join("");
}

function chart(elId, hintId, key, rows, hint) {
  const agg = {};
  for (const r of rows) { const k = r[key] || "\u2014 none"; agg[k] = (agg[k] || 0) + 1; }
  const items = Object.entries(agg).sort((a, b) => b[1] - a[1]);
  const max = Math.max(...items.map(i => i[1]), 1);
  document.getElementById(elId).innerHTML = items.length ? items.map(([name, n]) => {
    const color = key === "mkt" ? "var(" + mktVar(name) + ", var(--accent))" : "var(--accent)";
    const pct = (n / max * 100).toFixed(1);
    return '<div class="bar-row"><div class="name" title="' + esc(name) + '">' + esc(name) + '</div>' +
      '<div class="bar-track"><div class="bar-fill" style="width:' + pct + '%;background:' + color + '"></div></div>' +
      '<div class="amt">' + num(n) + '</div></div>';
  }).join("") : '<div class="bar-row"><div class="name" style="grid-column:1/-1;color:var(--muted)">Nothing matches the current filter</div></div>';
  document.getElementById(hintId).textContent = hint;
}

const uniq = key => [...new Set(DATA.map(r => r[key]).filter(Boolean))].sort();
const fMkt = document.getElementById("f-mkt"), fMonth = document.getElementById("f-month"),
      fCar = document.getElementById("f-car"), fStatus = document.getElementById("f-status"),
      fTrack = document.getElementById("f-track"), q = document.getElementById("q");
uniq("mkt").forEach(m => fMkt.add(new Option(m, m)));
uniq("car").forEach(c => fCar.add(new Option(c, c)));
uniq("st").forEach(s => fStatus.add(new Option(s, s)));
[...new Set(DATA.map(r => (r.shipIso || r.orderIso).slice(0, 7)).filter(Boolean))].sort().reverse()
  .forEach(ym => fMonth.add(new Option(fmtMonth(ym), ym)));

let sortKey = "ship", sortDir = -1;

function filtered() {
  const term = q.value.trim().toLowerCase();
  const fm = fMkt.value, fmo = fMonth.value, fc = fCar.value, fs = fStatus.value, ft = fTrack.value;
  const rows = DATA.filter(r => {
    if (fm && r.mkt !== fm) return false;
    if (fc && r.car !== fc) return false;
    if (fs && r.st !== fs) return false;
    if (ft === "yes" && !r.trk) return false;
    if (ft === "no" && r.trk) return false;
    if (fmo && (r.shipIso || r.orderIso).slice(0, 7) !== fmo) return false;
    if (term && !r.hay.includes(term)) return false;
    return true;
  });
  rows.sort((a, b) => {
    if (sortKey === "ship" || sortKey === "order" || sortKey === "units" || sortKey === "value" || sortKey === "cost") {
      // Rows with no ship date sort to the bottom either way &mdash; an empty cell is
      // not "oldest", it's unknown, and burying it keeps the top of the table useful.
      // An unpriced shipment sorts to the bottom in both directions, like a
      // missing ship date: "no invoice yet" is unknown, not cheap.
      if (sortKey === "cost") {
        if (a.cost == null && b.cost == null) return 0;
        if (a.cost == null) return 1;
        if (b.cost == null) return -1;
        return (a.cost - b.cost) * sortDir;
      }
      const x = a[sortKey], y = b[sortKey];
      if (sortKey !== "units" && sortKey !== "value") {
        if (x < 0 && y < 0) return 0;
        if (x < 0) return 1;
        if (y < 0) return -1;
      }
      return (x - y) * sortDir;
    }
    return String(a[sortKey]).localeCompare(String(b[sortKey])) * sortDir;
  });
  return rows;
}

// The table can hold tens of thousands of rows; paint a page at a time so the
// first render stays instant and the browser never builds 40k <tr> nodes at once.
const PAGE = 250;
let shown = PAGE, current = [];
// Keyed by the row's index in DATA, so a sort or filter in between doesn't move
// which order is open.
const expanded = new Set();

function lineRows(r) {
  if (!r.items.length) return '<p class="vnone" style="margin:0">No line detail on this order.</p>' +
    (r.cost == null ? "" : '<p class="linecost">Carrier charge: <b>' + money(r.cost) + '</b></p>');
  const rows = r.items.map(l => {
    const p = RAW.products[l[L.P]] || ["", ""];
    return '<tr><td class="sku">' + esc(p[P.SKU] || "\u2014") + '</td>' +
      '<td class="prod">' + esc(p[P.NAME]) + '</td>' +
      '<td class="num">' + num(l[L.QTY]) + '</td>' +
      '<td class="num">' + money(l[L.UNIT]) + '</td>' +
      '<td class="num">' + money(l[L.TOTAL]) + '</td></tr>';
  }).join("");
  const costLine = r.cost == null ? "" :
    '<p class="linecost">Carrier charge: <b>' + money(r.cost) + '</b>' +
    (r.csrc ? ' (' + esc(r.csrc) + ')' : "") +
    (r.units ? ' &middot; ' + money(r.cost / r.units) + ' per unit' : "") + '</p>';
  return '<table class="lines"><thead><tr><th>SKU</th><th>Item</th>' +
    '<th class="num">Qty</th><th class="num">Unit</th><th class="num">Line total</th></tr></thead>' +
    '<tbody>' + rows + '</tbody>' +
    (r.items.length > 1 ? '<tfoot><tr><td colspan="2"></td><td class="num">' + num(r.units) +
      '</td><td></td><td class="num">' + money(r.value) + '</td></tr></tfoot>' : "") +
    '</table>' + costLine;
}

function rowHtml(r) {
  const mkt = '<span class="chip" style="background:color-mix(in srgb, var(' + mktVar(r.mkt) +
    ') 15%, transparent);color:var(' + mktVar(r.mkt) + ')">' + esc(r.mkt) + '</span>';
  const stColor = r.st === "Shipped" ? "var(--ok)" : r.st === "Cancelled" ? "var(--dash)" :
                  r.st === "Open" || r.st === "Shipping" ? "var(--warn)" : "var(--muted)";
  const status = '<span class="chip" style="background:color-mix(in srgb, ' + stColor +
    ' 15%, transparent);color:' + stColor + '">' + esc(r.st) + '</span>';
  // Route the link off the tracking number's own shape first: the carrier field on
  // a marketplace feed is the requested carrier and is sometimes not the one that
  // actually carried it, and a wrong link is worse than a plain number.
  const base = TRACK_URL[carrierOf(r)];
  const trk = !r.trk ? '<span class="dash">&mdash;</span>'
    : base ? '<a class="link" href="' + base + encodeURIComponent(r.trk) + '" target="_blank" rel="noopener">' + esc(r.trk) + '</a>'
    : esc(r.trk);
  const open = expanded.has(r.i);
  const order = '<button class="expbtn" type="button" data-i="' + r.i + '" aria-expanded="' + open + '"' +
    ' title="Show the line items on this order"><span class="ecaret">&#9654;</span>' + esc(r.num) + '</button>';
  const main = '<tr>' +
    '<td class="date">' + (r.shipIso ? fmtDate(r.shipIso) : '<span class="dash">&mdash;</span>') + '</td>' +
    '<td class="date">' + (r.orderIso ? fmtDate(r.orderIso) : '<span class="dash">&mdash;</span>') + '</td>' +
    '<td>' + mkt + '</td>' +
    '<td class="ord">' + order + '</td>' +
    '<td class="cust">' + (r.cust ? esc(r.cust) : '<span class="dash" title="This marketplace redacts the customer name after the order ages out">&mdash;</span>') + '</td>' +
    '<td class="dest">' + (r.dest ? esc(r.dest) : '<span class="dash">&mdash;</span>') + '</td>' +
    '<td class="num">' + num(r.units) + (r.skus > 1 ? ' <small style="color:var(--muted)">/' + r.skus + '</small>' : "") + '</td>' +
    '<td class="num">' + (r.value ? money(r.value) : '<span class="dash">&mdash;</span>') + '</td>' +
    '<td class="num">' + (r.cost == null
      ? '<span class="dash" title="No carrier invoice matched to this tracking number yet">&mdash;</span>'
      : '<span title="' + esc(r.csrc || "carrier") + ' invoice">' + money(r.cost) + '</span>') + '</td>' +
    '<td>' + (r.car ? esc(r.car) : '<span class="dash">&mdash;</span>') + '</td>' +
    '<td>' + trk + '</td>' +
    '<td>' + status + '</td></tr>';
  return open
    ? main + '<tr class="linedetail"><td colspan="12">' + lineRows(r) + '</td></tr>'
    : main;
}

function paint(reset) {
  if (reset) { current = filtered(); shown = PAGE; }
  document.getElementById("rows").innerHTML =
    current.slice(0, shown).map(rowHtml).join("") ||
    '<tr><td colspan="12" style="color:var(--muted);padding:22px 12px">No shipments match the current filter.</td></tr>';
  const left = current.length - shown;
  const wrap = document.getElementById("morewrap");
  wrap.hidden = left <= 0;
  document.getElementById("more").textContent = "Show " + num(Math.min(left, 1000)) + " more";
  document.getElementById("count").textContent =
    num(current.length) + " shipment" + (current.length === 1 ? "" : "s") +
    (left > 0 ? " \u00b7 showing " + num(shown) : "");
  if (reset) {
    kpiCards(current);
    chart("chart-mkt", "hint-mkt", "mkt", current, "Shipments in the current filter");
    chart("chart-car", "hint-car", "car", current, "Carrier as recorded on the shipment");
  }
}

document.getElementById("rows").addEventListener("click", e => {
  const btn = e.target.closest(".expbtn");
  if (!btn) return;
  const i = +btn.dataset.i;
  if (expanded.has(i)) expanded.delete(i); else expanded.add(i);
  paint(false);
});

document.getElementById("more").addEventListener("click", () => { shown += 1000; paint(false); });
[q, fMkt, fMonth, fCar, fStatus, fTrack].forEach(el =>
  el.addEventListener("input", () => paint(true)));

document.querySelectorAll(".mp-table thead th[data-k]").forEach(th => {
  const go = () => {
    const k = th.dataset.k;
    sortDir = sortKey === k ? -sortDir
      : (k === "ship" || k === "order" || k === "units" || k === "value" || k === "cost" ? -1 : 1);
    sortKey = k;
    document.querySelectorAll(".mp-table thead th").forEach(o => o.classList.remove("sorted"));
    th.classList.add("sorted");
    th.querySelector(".arrow").textContent = sortDir === 1 ? "\u25b2" : "\u25bc";
    paint(true);
  };
  th.addEventListener("click", go);
  th.addEventListener("keydown", e => { if (e.key === "Enter" || e.key === " ") { e.preventDefault(); go(); } });
});

// The toolbar wraps to two rows on a narrow window, so measure it rather than
// hardcoding the offset the sticky headers park at.
function measureToolbar() {
  const bar = document.querySelector(".mp-toolbar");
  const h = bar.offsetHeight;
  const tall = h > window.innerHeight * 0.32;
  bar.classList.toggle("nostick", tall);
  document.documentElement.style.setProperty("--toolbar-h", tall ? "0px" : h + "px");
}
addEventListener("resize", measureToolbar);

// "Most rows are missing", not "all": once the warehouse reports fill some in,
// a strict zero test would silently drop the caveat while the column stays mostly empty.
const gaps = Object.entries(KPI.by_marketplace)
  .filter(([, v]) => v.n > 0 && v.shipdate / v.n < 0.6)
  .map(([k]) => k);
const costNote = KPI.cost_lo
  ? " <b>Ship cost</b> is the carrier's own invoice, matched by tracking number, and this build covers shipments from " +
    fmtDate(KPI.cost_lo) + " to " + fmtDate(KPI.cost_hi) + ". " +
    "FedEx bills weeks behind, so its recent shipments genuinely have no invoice yet. " +
    "<b>Stamps.com print history is available the same day</b> &mdash; a Stamps shipment with no cost means its export " +
    "has not been loaded into this build, not that the charge is pending."
  : " <b>Ship cost</b> is empty: no carrier invoices were loaded into this build.";
document.getElementById("gapnote").innerHTML = (gaps.length
  ? "<b>Known gap:</b> most " + gaps.join(" and ") + " rows carry no ship date or tracking number. " +
    "Those orders come from ShipStation, whose <i>shipment</i> feed stopped loading into BigQuery in October 2023 &mdash; " +
    "the order, customer and destination are current, the label is not. Target and Macy's are complete. " +
    "With no tracking number there is nothing for a carrier invoice to match, so those rows carry no ship cost either. " +
    "The warehouse shipped-order reports do have the tracking numbers, and this build recovered " +
    KPI.filled_3pl + " of them that way."
  : "") + costNote;

document.getElementById("foot").innerHTML =
  "<p>Built " + esc(KPI.built) + " from <code>acenda</code>, <code>macys</code> and <code>shipstation</code> in BigQuery &mdash; " +
  "a rolling " + KPI.days + "-day window.</p>" +
  "<p>Refresh with <code>python3 refresh_marketplace_shipments.py</code> in americanflat-ops-director.</p>";

document.querySelector(".mp-table thead th[data-k='ship']").classList.add("sorted");
measureToolbar();
paint(true);
</script>
"""


def render(rows, kpi):
    return (TEMPLATE
            .replace("__DATA__", json.dumps(encode(rows), separators=(",", ":")))
            .replace("__KPI__", json.dumps(kpi, separators=(",", ":"))))


# --------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--days", type=int, default=180,
                    help="rolling window in days (default 180)")
    ap.add_argument("--out", default="marketplace_shipments.html",
                    help="HTML output path")
    ap.add_argument("--ndjson", help="also write newline-delimited JSON for a BigQuery load")
    ap.add_argument("--3pl", nargs="*", default=[], metavar="PATH", dest="threepl",
                    help="warehouse shipped-order reports (NJ / Fontana / South "
                         "Carolina) to supply the ship date and tracking number "
                         "ShipStation no longer sends for Michaels and Shopify")
    ap.add_argument("--costs", nargs="*", default=[], metavar="PATH",
                    help="FedEx / Stamps.com invoice exports (files or a folder) to "
                         "price the shipments by tracking number")
    ap.add_argument("--order-costs", nargs="*", default=[], metavar="PATH",
                    help="the weekly shipping cost report's per-order roll-up "
                         "(all_orders_shipping_costs_*.md), joined by order number "
                         "to price orders that have no tracking number")
    ap.add_argument("--costs-ndjson", metavar="PATH",
                    help="write the parsed, de-duplicated charges as newline-delimited "
                         "JSON for loading into marketplaces.parcel_charges")
    ap.add_argument("--cost-table", metavar="TABLE",
                    help="BigQuery table of parcel charges (tracking, amount, carrier), "
                         "e.g. americanflat.marketplaces.parcel_charges")
    ap.add_argument("--source", choices=["feeds", "ledger"], default="feeds",
                    help="feeds = the live marketplace tables (default); "
                         "ledger = marketplaces.marketplace_shipments, once it exists")
    ap.add_argument("--auth", choices=["auto", "proxy", "gcloud"], default="auto",
                    help="auto = gcloud on the Mac, injected proxy creds in a cloud session")
    args = ap.parse_args()

    token = access_token(args.auth)
    print("Querying BigQuery for the last %d days…" % args.days, file=sys.stderr)
    rows = normalize(query(build_sql(args.days, args.source), token))

    filled = 0
    if args.threepl:
        print("Reading warehouse shipped-order reports:", file=sys.stderr)
        filled = attach_3pl(rows, load_3pl(args.threepl))
        print("  tracking filled in on %d shipments the marketplace feed left blank"
              % filled, file=sys.stderr)

    costs, charge_lines = {}, []
    if args.costs:
        print("Reading parcel charges:", file=sys.stderr)
        costs, charge_lines = load_costs(args.costs)
    if args.cost_table:
        for r in query(COST_TABLE_SQL % args.cost_table, token):
            costs[norm_tracking(r["tracking"])] = {"cost": float(r["cost"] or 0),
                                                   "source": r["source"] or "Other"}
    attach_costs(rows, costs)
    if args.costs_ndjson:
        today = dt.date.today().isoformat()
        with open(args.costs_ndjson, "w") as fh:
            for line in charge_lines:
                fh.write(json.dumps(dict(line, loaded_at=today)) + "\n")
        print("  %s: %d charge lines for BigQuery"
              % (args.costs_ndjson, len(charge_lines)), file=sys.stderr)
    if costs:
        priced = [r for r in rows if r["ship_cost"] is not None]
        units = sum(r["units"] for r in priced)
        print("  %d charges -> %d of %d shipments priced, $%s%s"
              % (len(costs), len(priced), len(rows),
                 format(round(sum(r["ship_cost"] for r in priced), 2), ",.2f"),
                 ", $%.2f per unit" % (sum(r["ship_cost"] for r in priced) / units)
                 if units else ""), file=sys.stderr)

    if args.order_costs:
        print("Reading the order-level shipping cost report:", file=sys.stderr)
        n = attach_order_costs(rows, load_order_costs(args.order_costs))
        print("  priced %d more shipments by order number" % n, file=sys.stderr)

    kpi = kpi_block(rows, args.days, filled)

    with open(args.out, "w") as fh:
        fh.write(render(rows, kpi))
    size = os.path.getsize(args.out)
    print("%s: %d shipments, %.1f MB" % (args.out, len(rows), size / 1e6), file=sys.stderr)

    if args.ndjson:
        loaded = dt.date.today().isoformat()
        with open(args.ndjson, "w") as fh:
            for r in rows:
                rec = dict(r, loaded_at=loaded, order_value=r["value"])
                rec.pop("value")
                if rec.get("ship_cost") is None:
                    rec.pop("ship_cost", None)
                # Empty strings are dropped so the MERGE sees NULL and keeps whatever
                # the ledger already holds, rather than blanking a captured fact.
                fh.write(json.dumps({k: v for k, v in rec.items() if v != ""}) + "\n")
        print("%s: %d rows for BigQuery" % (args.ndjson, len(rows)), file=sys.stderr)

    for m in MARKETPLACES:
        s = kpi["by_marketplace"].get(m)
        if not s:
            print("  %-9s no rows in window" % m, file=sys.stderr)
            continue
        print("  %-9s %6d shipments  %6d units  ship date %3d%%  tracking %3d%%"
              "  priced %3d%%"
              % (m, s["n"], s["units"], round(100 * s["shipdate"] / s["n"]),
                 round(100 * s["tracked"] / s["n"]),
                 round(100 * s["costed"] / s["n"])), file=sys.stderr)


if __name__ == "__main__":
    main()
